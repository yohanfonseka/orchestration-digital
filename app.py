# Orchestration-Digital v1.3 build=1787840016
import streamlit as st
import streamlit.components.v1 as components
import pandas as pd
import numpy as np
import json
import anthropic
from datetime import datetime, date
import io
from supabase import create_client, Client
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Orchestration-Digital", page_icon="🎯", layout="wide", initial_sidebar_state="expanded")

st.markdown("""
<style>
  @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&family=Plus+Jakarta+Sans:wght@400;500;600;700;800&display=swap');
  html, body, [class*="css"] { font-family:'Inter',sans-serif; background-color:#ffffff !important; color:#1a1d23; }
  .stApp { background:#f8f9fc !important; }

  /* ── Sidebar ── */
  [data-testid="stSidebar"] {
    background: linear-gradient(180deg, #1e3a5f 0%, #162d4a 100%) !important;
    border-right: none !important;
  }
  [data-testid="stSidebar"] * { color: #ffffff !important; }
  [data-testid="stSidebar"] .stRadio label {
    display: block !important;
    padding: 10px 16px !important;
    border-radius: 8px !important;
    font-size: 0.88rem !important;
    font-weight: 500 !important;
    color: rgba(255,255,255,0.75) !important;
    cursor: pointer !important;
    transition: all 0.15s !important;
    margin-bottom: 2px !important;
  }
  [data-testid="stSidebar"] .stRadio label:hover {
    background: rgba(255,255,255,0.1) !important;
    color: #ffffff !important;
  }
  [data-testid="stSidebar"] .stRadio [aria-checked="true"] + label,
  [data-testid="stSidebar"] .stRadio label:has(input:checked) {
    background: rgba(255,255,255,0.15) !important;
    color: #ffffff !important;
    font-weight: 600 !important;
  }
  [data-testid="stSidebar"] hr { border-color: rgba(255,255,255,0.15) !important; }
  [data-testid="stSidebar"] .stRadio > div { gap: 0 !important; }

  /* ── Hero ── */
  .hero-header { background:linear-gradient(135deg,#1e3a5f 0%,#2d5a9b 50%,#1a8a6e 100%); border-radius:16px; padding:36px 40px; margin-bottom:28px; }
  .hero-badge { display:inline-block; background:rgba(255,255,255,0.15); color:#fff; border:1px solid rgba(255,255,255,0.25); border-radius:20px; font-size:0.7rem; font-weight:600; letter-spacing:0.1em; text-transform:uppercase; padding:4px 14px; margin-bottom:12px; }
  .hero-title { font-family:'Plus Jakarta Sans',sans-serif; font-size:2rem; font-weight:800; color:#fff; letter-spacing:-0.5px; margin:0 0 8px 0; }
  .hero-sub { font-size:0.9rem; color:rgba(255,255,255,0.75); margin:0; }

  /* ── Cards ── */
  .metric-card { background:#fff; border:1px solid #e8eaf0; border-radius:12px; padding:20px 24px; text-align:center; box-shadow:0 1px 4px rgba(0,0,0,0.05); }
  .metric-value { font-family:'Plus Jakarta Sans',sans-serif; font-size:1.6rem; font-weight:700; color:#1e3a5f; }
  .metric-label { font-size:0.75rem; color:#8a93a8; text-transform:uppercase; letter-spacing:0.06em; margin-top:4px; }
  .section-header { font-family:'Plus Jakarta Sans',sans-serif; font-size:1rem; font-weight:700; color:#1a1d23; border-left:3px solid #2d5a9b; padding-left:12px; margin:24px 0 14px 0; }
  .settings-card { background:#fff; border:1px solid #e8eaf0; border-radius:12px; padding:20px 24px; margin-bottom:16px; box-shadow:0 1px 4px rgba(0,0,0,0.05); }

  /* ── Chat ── */
  .chat-container { background:#fff; border:1px solid #e8eaf0; border-radius:14px; overflow:hidden; box-shadow:0 1px 4px rgba(0,0,0,0.05); margin-bottom:16px; }
  .chat-header { background:linear-gradient(135deg,#1e3a5f,#2d5a9b); padding:16px 20px; }
  .chat-header-title { color:#fff; font-family:'Plus Jakarta Sans',sans-serif; font-weight:700; font-size:0.95rem; margin:0; }
  .chat-header-sub { color:rgba(255,255,255,0.7); font-size:0.78rem; margin:2px 0 0 0; }
  .chat-messages { padding:20px; max-height:520px; overflow-y:auto; scroll-behavior:smooth; }
  .msg-agent { display:flex; gap:12px; margin-bottom:16px; align-items:flex-start; }
  .msg-user { display:flex; gap:12px; margin-bottom:16px; align-items:flex-start; flex-direction:row-reverse; }
  .avatar-agent { width:32px; height:32px; background:linear-gradient(135deg,#1e3a5f,#2d5a9b); border-radius:50%; display:flex; align-items:center; justify-content:center; color:#fff; font-size:0.75rem; font-weight:700; flex-shrink:0; }
  .avatar-user { width:32px; height:32px; background:linear-gradient(135deg,#1a8a6e,#22b894); border-radius:50%; display:flex; align-items:center; justify-content:center; color:#fff; font-size:0.75rem; font-weight:700; flex-shrink:0; }
  .bubble-agent { background:#1e3a5f !important; border-radius:0 12px 12px 12px; padding:12px 16px; max-width:80%; font-size:0.88rem; line-height:1.6; color:#ffffff !important; border:1px solid #2d5a9b; }
  .bubble-user { background:linear-gradient(135deg,#1a8a6e,#22b894); border-radius:12px 0 12px 12px; padding:12px 16px; max-width:80%; font-size:0.88rem; line-height:1.6; color:#fff !important; }
  #chat-bottom { height:1px; }

  /* ── Buttons ── */
  .stButton > button { background:linear-gradient(135deg,#1e3a5f,#2d5a9b) !important; color:white !important; border:none !important; border-radius:8px !important; font-weight:600 !important; font-size:0.9rem !important; padding:10px 24px !important; }
  .stButton > button:hover { opacity:0.88 !important; }
  .green-btn .stButton > button { background:linear-gradient(135deg,#1a8a6e,#22b894) !important; }
  .plan-card { background:#fff; border:1px solid #e8eaf0; border-radius:14px; padding:28px 32px; margin-top:16px; line-height:1.8; color:#2d3341; box-shadow:0 1px 4px rgba(0,0,0,0.05); font-size:0.92rem; }

  /* ── Inputs ── */
  .stTextInput > div > div > input, .stNumberInput > div > div > input,
  .stTextArea textarea, .stDateInput > div > div > input {
    background:#fff !important; border:1px solid #d0d5e0 !important; border-radius:8px !important; color:#1a1d23 !important; font-size:0.9rem !important;
  }
  .stSelectbox > div > div { background:#fff !important; border:1px solid #d0d5e0 !important; border-radius:8px !important; color:#1a1d23 !important; }
  .stTabs [data-baseweb="tab-list"] { background:#eef1f8; border-radius:10px; padding:4px; gap:4px; }
  .stTabs [data-baseweb="tab"] { background:transparent !important; color:#6b7590 !important; border-radius:7px !important; font-weight:500 !important; font-size:0.88rem !important; }
  .stTabs [aria-selected="true"] { background:#fff !important; color:#1e3a5f !important; font-weight:600 !important; box-shadow:0 1px 4px rgba(0,0,0,0.08) !important; }
  div[data-testid="stExpander"] { background:#fff !important; border:1px solid #e8eaf0 !important; border-radius:10px !important; }
  label { color:#4a5168 !important; font-size:0.85rem !important; font-weight:500 !important; }
  #MainMenu, footer, header { visibility:hidden; }

  /* Force chat input to appear inline in the flow above action buttons */
  [data-testid="stChatInput"] {
    position: relative !important;
    bottom: auto !important;
    border-top: 1px solid #e8eaf0 !important;
    background: #ffffff !important;
    border-radius: 12px !important;
    margin: 12px 0 !important;
    box-shadow: 0 1px 4px rgba(0,0,0,0.06) !important;
  }
  [data-testid="stChatInput"] textarea {
    background: #ffffff !important;
    color: #1a1d23 !important;
    font-size: 0.9rem !important;
  }
  [data-testid="stChatInputSubmitButton"] svg {
    fill: #1e3a5f !important;
  }
</style>
""", unsafe_allow_html=True)

# Auto-scroll JS
AUTO_SCROLL_JS = """
<script>
  function scrollChat() {
    const el = document.getElementById('chat-bottom');
    if (el) el.scrollIntoView({ behavior: 'smooth' });
    const msgs = document.querySelector('.chat-messages');
    if (msgs) msgs.scrollTop = msgs.scrollHeight;
  }
  setTimeout(scrollChat, 120);
  setTimeout(scrollChat, 400);
</script>
"""

@st.cache_resource
def get_supabase():
    return create_client(st.secrets["SUPABASE_URL"], st.secrets["SUPABASE_KEY"])

def get_anthropic_client():
    return anthropic.Anthropic(api_key=st.secrets["ANTHROPIC_API_KEY"])

# ── Supabase helpers ──────────────────────────────────────────────────────────
def get_clients_list(sb):
    try: return sb.table("clients").select("*").order("client_name").execute().data or []
    except: return []

def get_brands_for_client(sb, client_id):
    try: return sb.table("brands").select("*").eq("client_id", client_id).order("brand_name").execute().data or []
    except: return []

def create_client_record(sb, name):
    try:
        r = sb.table("clients").insert({"client_name": name, "created_at": datetime.utcnow().isoformat()}).execute()
        return r.data[0] if r.data else None
    except Exception as e: st.error(f"Error: {e}"); return None

def create_brand_record(sb, client_id, name):
    try:
        r = sb.table("brands").insert({"client_id": client_id, "brand_name": name, "created_at": datetime.utcnow().isoformat()}).execute()
        return r.data[0] if r.data else None
    except Exception as e: st.error(f"Error: {e}"); return None

def get_brand_data_files(sb, brand_id):
    try: return sb.table("brand_data").select("*").eq("brand_id", brand_id).order("uploaded_at", desc=True).execute().data or []
    except: return []

def save_brand_data(sb, brand_id, file_name, data_json, row_count):
    try: sb.table("brand_data").insert({"brand_id": brand_id, "file_name": file_name, "data_json": data_json, "row_count": row_count, "uploaded_at": datetime.utcnow().isoformat()}).execute(); return True
    except Exception as e: st.error(f"Error: {e}"); return False

def delete_brand_data_file(sb, file_id):
    try: sb.table("brand_data").delete().eq("id", file_id).execute(); return True
    except: return False

def save_plan(sb, plan_data):
    try: sb.table("campaign_plans").insert(plan_data).execute(); return True
    except Exception as e: st.warning(f"Could not save: {e}"); return False

def delete_plan(sb, plan_id):
    try: sb.table("campaign_plans").delete().eq("id", plan_id).execute(); return True
    except Exception as e: st.warning(f"Could not delete: {e}"); return False

def load_saved_plans(sb):
    try: return sb.table("campaign_plans").select("*").order("created_at", desc=True).limit(50).execute().data or []
    except: return []

def get_client_benchmarks(sb, client_id):
    try: return sb.table("client_benchmarks").select("*").eq("client_id", str(client_id)).execute().data or []
    except: return []

def save_client_benchmark(sb, client_id, platform, data):
    try:
        existing = sb.table("client_benchmarks").select("id").eq("client_id", str(client_id)).eq("platform", platform).execute().data
        payload = {"client_id": str(client_id), "platform": platform, **data, "updated_at": datetime.utcnow().isoformat()}
        if existing: sb.table("client_benchmarks").update(payload).eq("client_id", str(client_id)).eq("platform", platform).execute()
        else: sb.table("client_benchmarks").insert(payload).execute()
        return True
    except Exception as e: st.error(f"Error saving benchmark: {e}"); return False

def get_client_benchmarks_as_dict(sb, client_id):
    """Returns benchmarks as {platform: {cpm, cpc, cpv, cpa, ctr, roas}} for use in planning."""
    rows = get_client_benchmarks(sb, client_id)
    result = {}
    for r in rows:
        result[r["platform"]] = {
            "cpm": r.get("cpm"), "cpc": r.get("cpc"), "cpv": r.get("cpv"),
            "cpa": r.get("cpa"), "ctr": r.get("ctr"), "roas": r.get("roas"),
            "rows": r.get("sample_size", 0), "is_manual": True
        }
    return result

def get_plan_version(sb, campaign_name, brand_id):
    try:
        plans = sb.table("campaign_plans").select("plan_version").eq("campaign_name", campaign_name).eq("brand_id", str(brand_id)).execute().data or []
        if not plans: return "V1.0"
        versions = []
        for p in plans:
            try: versions.append(float(p.get("plan_version","V1.0").replace("V","")))
            except: versions.append(1.0)
        return f"V{max(versions)+1:.1f}"
    except: return "V1.0"

def get_platform_settings(sb):
    try: return sb.table("platform_settings").select("*").order("platform_name").execute().data or []
    except: return []

def save_platform_setting(sb, platform_name, data):
    try:
        existing = sb.table("platform_settings").select("id").eq("platform_name", platform_name).execute().data
        if existing: sb.table("platform_settings").update({**data, "updated_at": datetime.utcnow().isoformat()}).eq("platform_name", platform_name).execute()
        else: sb.table("platform_settings").insert({"platform_name": platform_name, **data, "updated_at": datetime.utcnow().isoformat()}).execute()
        return True
    except Exception as e: st.error(f"Error: {e}"); return False

# ── Data helpers ──────────────────────────────────────────────────────────────
def parse_uploaded_file(f):
    n = f.name.lower()
    if n.endswith(".csv"): return pd.read_csv(f)
    elif n.endswith((".xlsx",".xls")): return pd.read_excel(f)
    return None

def extract_channel_benchmarks(df):
    benchmarks = {}
    col_lower = {c.lower(): c for c in df.columns}
    platform_keywords = {
        "Facebook":["facebook","fb"],"Instagram":["instagram","ig"],
        "TikTok":["tiktok","tik tok"],"YouTube":["youtube","yt"],
        "Google Search":["google search","search","sem"],
        "Google Display":["google display","display","gdn"],
        "LinkedIn":["linkedin"],"Programmatic":["programmatic","dsp"],
    }
    channel_col = next((c for cl,c in col_lower.items() if any(k in cl for k in ["channel","platform","campaign","source"])), None)

    def get_avg(sub_df, key):
        matches = [c for cl,c in {cc.lower():cc for cc in sub_df.columns}.items() if key in cl]
        if matches:
            try:
                vals = pd.to_numeric(sub_df[matches[0]], errors="coerce").dropna()
                return round(float(vals.mean()),2) if len(vals)>0 else None
            except: return None
        return None

    if channel_col:
        for platform, keywords in platform_keywords.items():
            mask = df[channel_col].astype(str).str.lower().apply(lambda x: any(kw in x for kw in keywords))
            sub = df[mask]
            if len(sub)>0:
                benchmarks[platform] = {
                    "cpm":get_avg(sub,"cpm"),"cpc":get_avg(sub,"cpc"),
                    "cpv":get_avg(sub,"cpv"),"cpa":get_avg(sub,"cpa"),
                    "ctr":get_avg(sub,"ctr"),"roas":get_avg(sub,"roas"),"rows":len(sub)
                }
    # Global fallback
    global_stats = {
        "cpm":get_avg(df,"cpm"),"cpc":get_avg(df,"cpc"),
        "cpv":get_avg(df,"cpv"),"cpa":get_avg(df,"cpa"),
        "ctr":get_avg(df,"ctr"),"roas":get_avg(df,"roas"),"rows":len(df)
    }
    for platform in platform_keywords:
        if platform not in benchmarks:
            benchmarks[platform] = {**global_stats, "is_global_fallback": True}
    return benchmarks

# Sri Lanka industry averages (LKR)
INDUSTRY_AVERAGES = {
    "Facebook":      {"cpm":420,"cpc":85,"cpv":None,"cpa":950,"ctr":1.2,"roas":2.8},
    "Instagram":     {"cpm":480,"cpc":95,"cpv":None,"cpa":1100,"ctr":1.0,"roas":2.5},
    "YouTube":       {"cpm":350,"cpc":None,"cpv":9,"cpa":None,"ctr":0.4,"roas":None},
    "Google Search": {"cpm":None,"cpc":75,"cpv":None,"cpa":820,"ctr":3.5,"roas":4.2},
    "Google Display":{"cpm":280,"cpc":60,"cpv":None,"cpa":1200,"ctr":0.35,"roas":None},
    "TikTok":        {"cpm":380,"cpc":90,"cpv":7,"cpa":1050,"ctr":1.5,"roas":2.2},
    "LinkedIn":      {"cpm":1200,"cpc":320,"cpv":None,"cpa":2500,"ctr":0.6,"roas":None},
    "Programmatic":  {"cpm":250,"cpc":55,"cpv":None,"cpa":1400,"ctr":0.25,"roas":None},
}

def calculate_budget_split(channels, total_budget, objective, audience_sizes, benchmarks, channel_kpis=None, channel_dates=None):
    """
    Optimised budget split:
    1. Calculate cost to reach 80% of targetable audience per channel
    2. Cap each channel at that 80% reach budget
    3. Redistribute surplus to most efficient channels by CPR
    4. Score remaining by objective fit + efficiency
    """
    objective_weights = {
        "Brand Awareness":         {"Facebook":1.3,"Instagram":1.2,"YouTube":1.4,"Google Display":1.0,"TikTok":1.1,"Google Search":0.6,"LinkedIn":0.8,"Programmatic":0.9},
        "Reach & Frequency":       {"Facebook":1.4,"Instagram":1.3,"YouTube":1.3,"Google Display":0.9,"TikTok":1.0,"Google Search":0.5,"LinkedIn":0.7,"Programmatic":0.8},
        "Video Views":             {"YouTube":1.5,"TikTok":1.4,"Facebook":1.1,"Instagram":1.2,"Google Display":0.8,"Google Search":0.4,"LinkedIn":0.7,"Programmatic":0.9},
        "Website Traffic":         {"Google Search":1.5,"Facebook":1.1,"Instagram":1.0,"TikTok":0.9,"Google Display":1.0,"YouTube":0.8,"LinkedIn":1.0,"Programmatic":0.9},
        "Lead Generation":         {"Google Search":1.4,"Facebook":1.3,"LinkedIn":1.5,"Instagram":1.1,"TikTok":0.8,"YouTube":0.7,"Google Display":0.9,"Programmatic":0.8},
        "E-commerce / Conversions":{"Google Search":1.5,"Facebook":1.3,"Instagram":1.2,"TikTok":1.0,"YouTube":0.9,"Google Display":1.0,"LinkedIn":0.7,"Programmatic":1.1},
        "App Installs":            {"Facebook":1.3,"Instagram":1.2,"TikTok":1.4,"Google Search":1.1,"YouTube":1.0,"Google Display":0.9,"LinkedIn":0.7,"Programmatic":0.9},
        "Engagement":              {"Instagram":1.4,"TikTok":1.5,"Facebook":1.2,"YouTube":1.0,"LinkedIn":0.9,"Google Search":0.5,"Google Display":0.7,"Programmatic":0.7},
    }
    obj_key = next((k for k in objective_weights if k.lower() in objective.lower()), "Brand Awareness")
    weights = objective_weights[obj_key]
    scores = {}
    total_budget = float(total_budget or 0)
    if not channels: return {}
    if channel_kpis is None: channel_kpis = {}

    REACH_CAP = 0.80      # 80% max reach rule
    AVG_FREQUENCY = 3.0   # assumed average frequency for reach calculation

    channel_data = {}
    for ch in channels:
        aud_raw = audience_sizes.get(ch, 500000)
        try: aud = int(float(str(aud_raw).replace(",","")))
        except: aud = 500000

        bench = benchmarks.get(ch, {})
        cpm_raw = bench.get("cpm") or INDUSTRY_AVERAGES.get(ch,{}).get("cpm",450)
        try: cpm = float(str(cpm_raw).replace(",",""))
        except: cpm = 450

        # Determine CPR based on KPI type
        kpi_type = channel_kpis.get(ch, "CPM")
        if "cpc" in kpi_type.lower():
            cpr_raw = bench.get("cpc") or INDUSTRY_AVERAGES.get(ch,{}).get("cpc",95)
        elif "cpv" in kpi_type.lower():
            cpr_raw = bench.get("cpv") or INDUSTRY_AVERAGES.get(ch,{}).get("cpv",10)
        elif "cpa" in kpi_type.lower() or "cpl" in kpi_type.lower():
            cpr_raw = bench.get("cpa") or INDUSTRY_AVERAGES.get(ch,{}).get("cpa",950)
        else:
            cpr_raw = cpm  # CPM is the cost per result for awareness

        try: cpr = float(str(cpr_raw).replace(",",""))
        except: cpr = 450

        # Budget needed to reach 80% of targetable audience
        reach_80pct = aud * REACH_CAP
        impressions_needed = reach_80pct * AVG_FREQUENCY
        budget_for_80pct = (impressions_needed / 1000) * cpm

        # Objective weight
        wt = next((v for k,v in weights.items() if k.lower() in ch.lower()), 1.0)

        channel_data[ch] = {
            "audience": aud,
            "cpm": cpm,
            "cpr": cpr,
            "efficiency": 1.0 / max(cpr, 0.01),  # lower CPR = higher efficiency
            "budget_80pct": budget_for_80pct,
            "weight": wt,
        }

    # Step 1: Initial allocation capped at 80% reach budget per channel
    total_80pct = sum(d["budget_80pct"] for d in channel_data.values())

    if total_80pct <= total_budget:
        # Budget exceeds all 80% reach costs — cap each and redistribute surplus
        split = {ch: min(d["budget_80pct"], total_budget) for ch,d in channel_data.items()}
        surplus = total_budget - sum(split.values())
        # Distribute surplus to most efficient channels (raise their reach % proportionally)
        if surplus > 0:
            efficiency_sum = sum(d["efficiency"] * d["weight"] for d in channel_data.values())
            for ch in channels:
                d = channel_data[ch]
                split[ch] += round(surplus * (d["efficiency"] * d["weight"]) / max(efficiency_sum, 0.001), 0)
    else:
        # Not enough budget to reach 80% everywhere — prioritise by CPR efficiency
        scores = {ch: d["efficiency"] * d["weight"] for ch,d in channel_data.items()}
        total_score = sum(scores.values()) or 1
        split = {ch: round((scores[ch]/total_score)*total_budget, 0) for ch in channels}

    # Enforce minimum 5% per channel
    min_budget = total_budget * 0.05
    for ch in split:
        if split[ch] < min_budget: split[ch] = min_budget

    # Re-normalise to total budget
    total_alloc = sum(split.values()) or 1
    return {ch: round(v/total_alloc*total_budget, 0) for ch,v in split.items()}

def calculate_channel_kpis(channel, budget_lkr, benchmarks, objective):
    try: budget_lkr = float(str(budget_lkr).replace(",",""))
    except: budget_lkr = 0
    bench = benchmarks.get(channel, {})
    industry = INDUSTRY_AVERAGES.get(channel, {})
    objective_kpi_map = {
        "awareness":("CPM","Impressions"),"reach":("CPM","Impressions"),
        "video":("CPV","Video Views"),"traffic":("CPC","Clicks"),
        "lead":("CPL","Leads"),"conversion":("CPA","Conversions"),
        "ecommerce":("CPA","Conversions"),"app":("CPC","Installs"),
        "engagement":("CPE","Engagements"),
    }
    obj_lower = objective.lower()
    kpi_type, kpi_metric = next(((v[0],v[1]) for k,v in objective_kpi_map.items() if k in obj_lower),("CPM","Impressions"))
    if "search" in channel.lower(): kpi_type,kpi_metric = "CPC","Clicks"
    elif "youtube" in channel.lower(): kpi_type,kpi_metric = "CPV","Video Views"
    elif "linkedin" in channel.lower() and "lead" in obj_lower: kpi_type,kpi_metric = "CPL","Leads"

    rate_key = kpi_type.lower().replace("cpl","cpa").replace("cpe","cpc")
    historical_rate = bench.get(rate_key)
    is_industry_avg = False
    data_source = "Historical Data"

    if not historical_rate or bench.get("is_global_fallback"):
        industry_rate = industry.get(rate_key)
        if industry_rate:
            historical_rate = industry_rate
            is_industry_avg = True
            data_source = "⚠️ Industry Average (LK)"
        else:
            fallback = {"cpm":450,"cpc":95,"cpv":12,"cpa":850}
            historical_rate = fallback.get(rate_key,450)
            is_industry_avg = True
            data_source = "⚠️ Industry Average (LK)"

    try: buying_rate_lkr = round(float(str(historical_rate).replace(",","")), 2)
    except: buying_rate_lkr = 450.0
    if buying_rate_lkr <= 0: buying_rate_lkr = 450.0
    if kpi_type=="CPM":
        target=int(budget_lkr/buying_rate_lkr*1000); target_str=f"{target:,} Impressions"; rate_str=f"LKR {buying_rate_lkr:,.0f} CPM"
    elif kpi_type=="CPC":
        target=int(budget_lkr/buying_rate_lkr); target_str=f"{target:,} Clicks"; rate_str=f"LKR {buying_rate_lkr:,.0f} CPC"
    elif kpi_type=="CPV":
        target=int(budget_lkr/buying_rate_lkr); target_str=f"{target:,} Video Views"; rate_str=f"LKR {buying_rate_lkr:,.2f} CPV"
    elif kpi_type in ("CPA","CPL"):
        target=int(budget_lkr/buying_rate_lkr); target_str=f"{target:,} {kpi_metric}"; rate_str=f"LKR {buying_rate_lkr:,.0f} {kpi_type}"
    elif kpi_type=="ROAS":
        roas=bench.get("roas") or industry.get("roas") or 3.0; target_str=f"{roas:.1f}x ROAS"; rate_str=f"{roas:.1f}x"
    else:
        target=int(budget_lkr/buying_rate_lkr*1000); target_str=f"{target:,} Impressions"; rate_str=f"LKR {buying_rate_lkr:,.0f}"

    ctr = bench.get("ctr") or (industry.get("ctr") if is_industry_avg else None)
    return {"kpi_type":kpi_type,"buying_rate":rate_str,"target_kpi":target_str,
            "ctr_bench":f"CTR: {ctr:.2f}%" if ctr else "","raw_rate":buying_rate_lkr,
            "is_industry_avg":is_industry_avg,"data_source":data_source}

def get_data_gaps(channels, benchmarks):
    gaps = []
    for ch in channels:
        bench = benchmarks.get(ch,{})
        if not bench or bench.get("is_global_fallback") or (not bench.get("cpm") and not bench.get("cpc")):
            ia = INDUSTRY_AVERAGES.get(ch,{})
            gaps.append(f"**{ch}**: No historical data found — using Sri Lanka industry average "
                       f"(CPM: LKR {ia.get('cpm','—')}, CPC: LKR {ia.get('cpc','—')})")
    return gaps

def summarise_dataframe(df):
    """Compact summary — benchmarks only, no raw data rows. Saves ~60% tokens."""
    col_lower = {c.lower():c for c in df.columns}
    lines = [f"Dataset: {len(df):,} rows, {len(df.columns)} columns",
             f"Columns: {', '.join(df.columns.tolist())}","","Performance benchmarks:"]
    for metric in ["cpm","cpc","ctr","roas","cpa","cpv","spend","impressions","clicks","conversions","reach","frequency"]:
        matches = [c for cl,c in col_lower.items() if metric in cl]
        if matches:
            try:
                vals = pd.to_numeric(df[matches[0]],errors="coerce").dropna()
                if len(vals)>0:
                    lines.append(f"  {matches[0]}: avg={vals.mean():.2f}, median={vals.median():.2f}, min={vals.min():.2f}, max={vals.max():.2f} (n={len(vals)})")
            except: pass
    # Channel breakdown if available
    ch_col = next((c for cl,c in col_lower.items() if any(k in cl for k in ["channel","platform","campaign name","source"])), None)
    if ch_col:
        lines += ["","Channel breakdown:"]
        for ch_val in df[ch_col].dropna().unique()[:12]:
            sub = df[df[ch_col]==ch_val]
            sub_metrics = []
            for m in ["cpm","cpc","ctr","roas"]:
                sub_matches = [c for cl,c in {cc.lower():cc for cc in sub.columns}.items() if m in cl]
                if sub_matches:
                    try:
                        v = pd.to_numeric(sub[sub_matches[0]],errors="coerce").dropna()
                        if len(v)>0: sub_metrics.append(f"{m}={v.mean():.2f}")
                    except: pass
            if sub_metrics:
                lines.append(f"  {ch_val} ({len(sub)} rows): {', '.join(sub_metrics)}")
    return "\n".join(lines)

def format_lkr(v):
    try: return f"LKR {float(v):,.0f}"
    except: return "LKR 0"

# ── Chat ──────────────────────────────────────────────────────────────────────
def render_message(role, content):
    avatar = "AI" if role=="agent" else "You"
    avatar_cls = "avatar-agent" if role=="agent" else "avatar-user"
    bubble_cls = "bubble-agent" if role=="agent" else "bubble-user"
    msg_cls = "msg-agent" if role=="agent" else "msg-user"
    st.markdown(f"""<div class="{msg_cls}"><div class="{avatar_cls}">{avatar}</div><div class="{bubble_cls}">{content.replace(chr(10),'<br>')}</div></div>""", unsafe_allow_html=True)

PLANNING_SYSTEM = """You are Orchy, an expert digital media planning agent inside Orchestration-Digital, used by planners in Sri Lanka.

Collect everything needed for a complete digital media plan through friendly, professional conversation.

CONVERSATION FLOW — cover in order, 1–2 questions at a time:
1. Campaign name and objective (Awareness / Traffic / Leads / Conversions / App Installs / Engagement / Video Views)
2. Total budget in LKR and flight dates
3. Target audience description (demographics, interests, behaviours)
4. Channels — ask about each: Facebook, Instagram, YouTube, Google Search, Google Display, TikTok, LinkedIn, Programmatic Display
5. For each selected channel, ask for the TARGETABLE AUDIENCE SIZE from the platform estimator (Meta Audience Insights, Google Reach Planner, TikTok Audience Estimator). Be specific about where to find this.
6. For each channel, ask what creative assets are available: Static Images, Videos, Carousels, Stories, UGC, Influencer Content
   - For each asset type, confirm: placement, objective, and KPI
   - Ask: "Would you like to allocate separate budgets per creative type, or keep the total budget at platform level?"
   - If per-creative: ask for the budget split across creative types for that channel
7. Any additional context or constraints

RULES:
- Conversational and professional, 1–2 questions at a time
- Confirm back what you've heard
- When ALL info collected, end with exactly: [BRIEF_COMPLETE]
- Never generate the plan yourself
- All budgets in LKR
- When asking for audience sizes, explain exactly where to find them on each platform"""

EDIT_SYSTEM = """You are Orchy, an expert digital media planning agent. You are helping a planner edit an existing campaign plan.

You have access to:
1. The original planning conversation that created this plan (so you understand WHY decisions were made)
2. The full plan that was generated from that conversation

Your job is to:
1. Use the original conversation context to understand the strategic intent behind the plan
2. Understand what changes the planner wants to make
3. Ask clarifying questions if needed
4. When you have enough information, confirm the changes and end with exactly: [EDIT_COMPLETE]

Be specific — confirm budget numbers, channel changes, and KPI impacts before finalising.
All budgets in LKR."""

def get_agent_response(messages, client_name, brand_name, data_summary, mode="planning", existing_plan="", original_conversation=""):
    client = get_anthropic_client()
    if mode=="editing":
        conv_context = f"\n\nORIGINAL PLANNING CONVERSATION (context for why decisions were made):\n{original_conversation[:3000]}" if original_conversation else ""
        system = EDIT_SYSTEM + f"\n\nCLIENT: {client_name}\nBRAND: {brand_name}{conv_context}\n\nORIGINAL PLAN:\n{existing_plan[:2000]}"
    else:
        system = PLANNING_SYSTEM + f"\n\nCLIENT: {client_name}\nBRAND: {brand_name}\n\nHISTORICAL BENCHMARKS:\n{data_summary[:1500]}"

    # Send full conversation for planning — Orchy needs all context to plan correctly.
    # For very long chats (20+ msgs), summarise only the oldest messages to save tokens
    # while keeping recent context intact.
    if len(messages) <= 20:
        trimmed = messages
    else:
        # Summarise earliest messages into a compact brief, keep last 16 in full
        early = messages[:-16]
        recent = messages[-16:]
        # Build a compact summary of early messages
        early_text = " | ".join([
            f"{'P' if m['role']=='user' else 'A'}: {m['content'][:120].replace(chr(10),' ')}"
            for m in early
        ])
        summary_note = f"[Summary of earlier discussion: {early_text}]"
        trimmed = [{"role":"user","content":summary_note}] + recent

    api_messages = [{"role":m["role"].replace("agent","assistant"),"content":m["content"]} for m in trimmed]
    response = client.messages.create(model="claude-sonnet-4-6",max_tokens=1000,system=system,messages=api_messages)
    return response.content[0].text

def extract_brief_from_conversation(conversation):
    """Extract brief JSON from conversation — uses last 12 messages only to save tokens."""
    client = get_anthropic_client()
    # Use last 12 messages (where all key info will be)
    trimmed = conversation[-12:] if len(conversation) > 12 else conversation
    conv_text = "\n".join([f"{'PLANNER' if m['role']=='user' else 'AGENT'}: {m['content']}" for m in trimmed])
    prompt = f"""Extract the campaign brief. Return ONLY valid JSON with this exact structure:
{{
  "campaign_name": "",
  "objective": "",
  "total_budget": 0,
  "start_date": "YYYY-MM-DD",
  "end_date": "YYYY-MM-DD",
  "audience": "",
  "channels": [],
  "audience_sizes": {{}},
  "assets": "",
  "market": "Sri Lanka",
  "channel_placements": {{
    "ChannelName": [
      {{"placement": "Feed", "kpi_type": "CPM", "assets": "Static Image", "objective": "Awareness", "budget": 0}}
    ]
  }}
}}

Notes:
- channel_placements: only populate if the planner specified per-creative budget splits. If platform-level only, leave as empty object {{}}.
- budget in channel_placements is in LKR. Sum of placements for a channel should equal that channel's total budget.
- If no per-creative split was discussed, leave channel_placements as {{}}.

CONVERSATION:
{conv_text}

JSON only, no markdown."""
    msg = client.messages.create(model="claude-sonnet-4-6",max_tokens=600,messages=[{"role":"user","content":prompt}])
    try:
        text=msg.content[0].text.strip().replace("```json","").replace("```","").strip()
        return json.loads(text)
    except:
        return {"campaign_name":"Campaign","objective":"","total_budget":0,"start_date":str(date.today()),
                "end_date":str(date.today()),"audience":"","channels":[],"audience_sizes":{},"assets":"","market":"Sri Lanka"}

def generate_media_plan(brief, client_name, brand_name, data_summary, budget_split, channel_kpi_data, data_gaps):
    """Generate full media plan from structured brief form data + pre-calculated budget/KPI data."""
    client = get_anthropic_client()
    # Build structured brief text from form data
    conv_text = f"""
CAMPAIGN NAME: {brief.get("campaign_name","")}
OBJECTIVE: {brief.get("objective","")}
TOTAL BUDGET: LKR {brief.get("total_budget",0):,.0f}
FLIGHT DATES: {brief.get("start_date","")} to {brief.get("end_date","")}
MARKET: {brief.get("market","Sri Lanka")}
BUDGET TYPE: {brief.get("budget_type","Total campaign budget")}

TARGET AUDIENCE:
- Age: {brief.get("age_min",18)}–{brief.get("age_max",65)}
- Gender: {brief.get("gender","All")}
- Locations: {brief.get("locations","")}
- Interests & Behaviours: {brief.get("interests","")}
- Custom Audiences: {brief.get("custom_audiences","")}
- Lookalike Audiences: {brief.get("lookalike","")}

CHANNELS & AUDIENCE SIZES:
{chr(10).join([f"  {ch}: targetable audience = {brief.get('audience_sizes',{}).get(ch,'unknown'):,} | KPI = {brief.get('channel_kpis',{}).get(ch,'CPM')} | Dates: {brief.get('channel_dates',{}).get(ch,'same as campaign')}" for ch in brief.get("channels",[])])}

CREATIVE ASSETS PER CHANNEL:
{chr(10).join([f"  {ch}: {', '.join(brief.get('creative_assets',{}).get(ch,[]))}" for ch in brief.get("channels",[])])}

ADDITIONAL PLANNING INSTRUCTIONS:
{brief.get("planning_instructions","")}
"""
    split_lines = [
        f"  {ch}: LKR {b:,.0f} | KPI: {channel_kpi_data.get(ch,{}).get('kpi_type','CPM')} | "
        f"Buying Rate: {channel_kpi_data.get(ch,{}).get('buying_rate','—')} | "
        f"Target: {channel_kpi_data.get(ch,{}).get('target_kpi','—')} | "
        f"Source: {channel_kpi_data.get(ch,{}).get('data_source','—')}"
        for ch,b in budget_split.items()
    ]
    gaps_note = ("\n\nDATA GAPS — industry averages used for: " + ", ".join([g.split(":")[0].replace("**","").strip() for g in data_gaps])) if data_gaps else ""

    prompt = f"""You are a senior digital media planner with 15+ years of experience planning campaigns in Sri Lanka.

Generate a complete, professional media plan based on the campaign brief and pre-calculated data below.

CLIENT: {client_name}
BRAND: {brand_name}

CAMPAIGN BRIEF:
{conv_text}

PRE-CALCULATED BUDGET SPLIT & KPIs — use these exact numbers, do not recalculate:
{chr(10).join(split_lines)}{gaps_note}

PLANNING RULES (follow strictly):
1. Budget is optimised by minimum cost per result — channels with lower CPR get higher budget priority
2. 80% reach cap — never allocate budget beyond what is needed to reach 80% of the targetable audience on any channel (diminishing returns beyond this point)
3. If a channel hits 80% reach cap, redistribute surplus budget to the next most efficient channel
4. Prioritise channels in order: lowest CPR → largest reachable audience → objective fit
5. For awareness campaigns: prioritise lowest CPM channels first
6. For conversion campaigns: prioritise lowest CPA/CPL channels first
7. All buying rates and KPI targets are pre-calculated from historical data — use them exactly
8. Flag clearly where industry averages were used (no historical data)

HISTORICAL PERFORMANCE DATA:
{data_summary}

Generate the media plan with these sections. Use markdown tables where appropriate:

## 1. EXECUTIVE SUMMARY
Strategic rationale (3-4 sentences) referencing the brief, objective, and historical performance.

## 2. BUDGET ALLOCATION RATIONALE
Explain why each channel received its budget — reference audience sizes, historical efficiency, and objective fit.

## 3. CHANNEL-BY-CHANNEL PLAN
For each channel, show a table:
| Metric | Value |
|--------|-------|
| Budget (LKR) | ... |
| KPI Type | ... |
| Buying Rate | ... |
| Target KPI | ... |
| Ad Formats | ... |
| Targeting | ... |
| Creative Assets | ... |
| Data Source | ... |
Then 2-3 sentences on tactics, bidding strategy and pacing.

## 4. CREATIVE ASSET PLAN
| Asset Type | Platform | Placement | Objective | KPI Target | Specs |
|------------|----------|-----------|-----------|------------|-------|

## 5. REACH & FREQUENCY PROJECTIONS
Based on audience sizes and budgets — estimated reach % and avg frequency per platform.

## 6. KPI SUMMARY TABLE
| Channel | Budget (LKR) | KPI Type | Buying Rate | Target KPI | Data Source |
|---------|-------------|----------|-------------|------------|-------------|

## 7. WEEKLY OPTIMISATION ROADMAP
| Week | Focus | Key Actions |
|------|-------|-------------|

## 8. RISK FLAGS & MITIGATION
Bullet points — risks from historical data and mitigation strategies.

Use LKR throughout. Flag ⚠️ clearly where industry averages were used instead of historical data."""

    msg = client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=5000,
        messages=[{"role":"user","content":prompt}]
    )
    return msg.content[0].text

def apply_plan_edits(edit_conversation, original_plan, client_name, brand_name, budget_split, channel_kpi_data, original_conversation=""):
    client = get_anthropic_client()
    conv_text = "\n".join([f"{'PLANNER' if m['role']=='user' else 'AGENT'}: {m['content']}" for m in edit_conversation])
    split_lines = [f"  {ch}: LKR {b:,.0f} | {channel_kpi_data.get(ch,{}).get('kpi_type','CPM')} | Rate: {channel_kpi_data.get(ch,{}).get('buying_rate','—')} | Target: {channel_kpi_data.get(ch,{}).get('target_kpi','—')}" for ch,b in budget_split.items()]
    original_context = f"\n\nORIGINAL PLANNING CONVERSATION (strategic context):\n{original_conversation[:2000]}" if original_conversation else ""
    prompt = f"""You are a senior digital media planner. Apply the requested edits to this media plan.

CLIENT: {client_name} | BRAND: {brand_name}
{original_context}

ORIGINAL BUDGET SPLIT:
{chr(10).join(split_lines)}

EDIT INSTRUCTIONS FROM PLANNER:
{conv_text}

ORIGINAL PLAN:
{original_plan}

Produce the complete updated media plan with all edits applied. Keep the same structure but update all affected numbers, budgets, KPIs, and rationale. Clearly note what was changed at the top under "## CHANGES FROM PREVIOUS VERSION"."""

    msg = client.messages.create(model="claude-sonnet-4-6",max_tokens=5000,messages=[{"role":"user","content":prompt}])
    return msg.content[0].text

# ── Excel ──────────────────────────────────────────────────────────────────────
def build_excel(brief_summary, plan_text="", client_name="", brand_name="", budget_split={}, channel_kpi_data={}, plan_version="V1.0"):
    wb = Workbook(); ws = wb.active; ws.title = "Media Plan"
    navy="1E3A5F"; white="FFFFFF"; light="F0F4FA"; border_col="D0D5E0"

    def cs(row,col,value="",bold=False,bg=None,fg="1A1D23",align="left",size=10,num_fmt=None):
        c=ws.cell(row=row,column=col,value=value)
        c.font=Font(name="Inter",bold=bold,color=fg,size=size)
        if bg: c.fill=PatternFill("solid",fgColor=bg)
        c.alignment=Alignment(horizontal=align,vertical="center",wrap_text=True)
        thin=Side(style="thin",color=border_col)
        c.border=Border(left=thin,right=thin,top=thin,bottom=thin)
        if num_fmt: c.number_format=num_fmt
        return c

    def ms(r1,c1,r2,c2,value="",bold=False,bg=None,fg="1A1D23",align="left",size=10):
        ws.merge_cells(start_row=r1,start_column=c1,end_row=r2,end_column=c2)
        c=ws.cell(row=r1,column=c1,value=value)
        c.font=Font(name="Inter",bold=bold,color=fg,size=size)
        if bg: c.fill=PatternFill("solid",fgColor=bg)
        c.alignment=Alignment(horizontal=align,vertical="center",wrap_text=True)
        return c

    for i,w in enumerate([24,30,20,14,18,20,16,18,18,22,12,12,10],1):
        ws.column_dimensions[get_column_letter(i)].width=w

    ws.row_dimensions[1].height=40
    ms(1,1,1,13,"ORCHESTRATION-DIGITAL  |  MEDIA PLAN",bold=True,bg=navy,fg=white,align="center",size=14)

    info=brief_summary or {}
    meta=[("Date",datetime.today().strftime("%d %b %Y")),("Client",client_name),("Brand",brand_name),
          ("Campaign Name",info.get("campaign_name","")),("Campaign Start",info.get("start_date","")),
          ("Campaign End",info.get("end_date","")),("Plan Version",plan_version)]
    for i,(label,val) in enumerate(meta,2):
        ws.row_dimensions[i].height=20
        cs(i,1,label,bold=True,bg=light,fg=navy,align="right")
        ms(i,2,i,6,value=str(val))
        for c in range(7,14): cs(i,c,bg=white)

    hdr_row=10; ws.row_dimensions[hdr_row].height=40
    hdrs=["Channel","Objective","Target Audience","KPI Type","Buying Rate (LKR)","Target KPI","Spendable (USD)","Spendable (LKR)","Billable (LKR)","Creative Assets","Start Date","End Date","Days"]
    for ci,h in enumerate(hdrs,1): cs(hdr_row,ci,h,bold=True,bg=navy,fg=white,align="center",size=9)

    channel_colours={"Facebook":"1877F2","Instagram":"E1306C","YouTube":"FF0000",
                     "Google Search":"1A8A6E","Google Display":"34A853","TikTok":"010101",
                     "LinkedIn":"0A66C2","Programmatic":"7B5EA7"}
    usd_rate=320; commission=0.10; ssc_rate=0.025641; vat_rate=0.18; wht_rate=0.163
    start_str=info.get("start_date",""); end_str=info.get("end_date","")
    try:
        s=datetime.strptime(start_str,"%Y-%m-%d"); e=datetime.strptime(end_str,"%Y-%m-%d"); days=(e-s).days
    except: days=0

    current_row=hdr_row+1; channel_totals={}

    # creative_placements: dict of ch -> list of {placement, kpi_type, buying_rate, target_kpi, assets, budget}
    # If planner specified per-creative budgets, channel_kpi_data[ch] may contain a "placements" list
    for ch,sub_lkr in budget_split.items():
        colour=next((v for k,v in channel_colours.items() if k.lower() in ch.lower()),navy)
        ws.row_dimensions[current_row].height=22
        ms(current_row,1,current_row,13,ch.upper(),bold=True,bg=colour,fg=white,align="left",size=10)
        current_row+=1

        kpi=channel_kpi_data.get(ch,{})
        src_note=" ⚠️" if kpi.get("is_industry_avg") else ""
        placements=kpi.get("placements",[])  # list of per-creative placement rows

        if placements:
            # Write one row per creative placement
            for pl in placements:
                pl_lkr=pl.get("budget",round(sub_lkr/len(placements),0))
                pl_usd=round(pl_lkr/usd_rate,2)
                pl_is_meta=any(m in ch.lower() for m in ["facebook","instagram"])
                pl_billable=round(pl_lkr*1.05,0) if pl_is_meta else pl_lkr
                row_bg=light if current_row%2==0 else white
                ws.row_dimensions[current_row].height=22
                data=[pl.get("placement",ch),pl.get("objective",info.get("objective","")),
                      info.get("audience",""),pl.get("kpi_type",kpi.get("kpi_type","CPM")),
                      pl.get("buying_rate",kpi.get("buying_rate","—"))+src_note,
                      pl.get("target_kpi",kpi.get("target_kpi","—")),
                      pl_usd,pl_lkr,pl_billable,pl.get("assets",info.get("assets","")),
                      start_str,end_str,days]
                for ci,val in enumerate(data,1):
                    fmt='#,##0.00' if ci==7 else ('#,##0' if ci in (8,9) else None)
                    cs(current_row,ci,val,bg=row_bg,align="center" if ci>3 else "left",num_fmt=fmt)
                current_row+=1
            channel_totals[ch]=sub_lkr
        else:
            # Single row for channel total
            meta_channels=["facebook","instagram"]
            is_meta=any(m in ch.lower() for m in meta_channels)
            sub_usd=round(sub_lkr/usd_rate,2); billable=round(sub_lkr*1.05,0) if is_meta else sub_lkr
            row_bg=light if current_row%2==0 else white
            ws.row_dimensions[current_row].height=22
            data=[ch,info.get("objective",""),info.get("audience",""),kpi.get("kpi_type","CPM"),
                  kpi.get("buying_rate","—")+src_note,kpi.get("target_kpi","—"),
                  sub_usd,sub_lkr,billable,info.get("assets",""),start_str,end_str,days]
            for ci,val in enumerate(data,1):
                fmt='#,##0.00' if ci==7 else ('#,##0' if ci in (8,9) else None)
                cs(current_row,ci,val,bg=row_bg,align="center" if ci>3 else "left",num_fmt=fmt)
            channel_totals[ch]=sub_lkr; current_row+=1

    current_row+=1
    # 5% markup applies to Meta (Facebook/Instagram) only — all other channels billable = spendable
    meta_ch = ["facebook","instagram"]
    billable_totals = {ch: (round(v*1.05,0) if any(m in ch.lower() for m in meta_ch) else v) for ch,v in channel_totals.items()}
    total_billable = sum(billable_totals.values())
    agency_comm=round(total_billable*commission,2)
    sub1=total_billable+agency_comm
    ssc=round(sub1*ssc_rate,2)
    sub2=sub1+ssc
    vat=round(sub2*vat_rate,2)

    # WHT only on YouTube and Google billable spend
    wht_channels=["youtube","google search","google display"]
    wht_base=sum(billable_totals.get(ch,v) for ch,v in channel_totals.items() if any(w in ch.lower() for w in wht_channels))
    wht=round(wht_base*wht_rate,2) if wht_base>0 else 0
    total_invest=sub2+vat

    summary=[("Total Working Investment — Billable (LKR)",total_billable),
             ("Agency Commission (10%)",agency_comm),
             ("Sub Total",sub1),
             ("SSC Levy (2.5641%)",ssc),
             ("Sub Total",sub2),
             ("VAT (18%)",vat)]
    if wht>0:
        summary.append((f"Withholding Tax 16.3% (YouTube/Google — on LKR {wht_base:,.0f})",wht))
    summary.append(("TOTAL INVESTMENT (LKR)",total_invest))
    for label,val in summary:
        ws.row_dimensions[current_row].height=22
        is_total="TOTAL INVESTMENT" in label; is_sub=label.startswith("Sub") or label.startswith("Total W")
        bg=navy if is_total else (light if is_sub else white); fg_col=white if is_total else "1A1D23"
        ms(current_row,1,current_row,7,label,bold=is_total or is_sub,bg=bg,fg=fg_col,align="right")
        cs(current_row,8,val,bold=is_total or is_sub,bg=bg,fg=fg_col,align="right",num_fmt='#,##0.00')
        for c in range(9,14): cs(current_row,c,bg=white)
        current_row+=1

    ws3=wb.create_sheet("KPI Summary")
    for i,w in enumerate([22,18,14,22,26,20],1): ws3.column_dimensions[get_column_letter(i)].width=w
    ws3.merge_cells("A1:F1")
    h=ws3.cell(row=1,column=1,value="KPI SUMMARY BY CHANNEL")
    h.font=Font(name="Inter",bold=True,size=12,color=navy); h.fill=PatternFill("solid",fgColor=light)
    for ci,hdr in enumerate(["Channel","Budget (LKR)","KPI Type","Buying Rate","Target KPI","Data Source"],1):
        c=ws3.cell(row=2,column=ci,value=hdr); c.font=Font(name="Inter",bold=True,color=white,size=9)
        c.fill=PatternFill("solid",fgColor=navy); c.alignment=Alignment(horizontal="center",vertical="center")
    for i,(ch,budget) in enumerate(budget_split.items(),3):
        kpi=channel_kpi_data.get(ch,{})
        for ci,val in enumerate([ch,budget,kpi.get("kpi_type","—"),kpi.get("buying_rate","—"),kpi.get("target_kpi","—"),kpi.get("data_source","—")],1):
            c=ws3.cell(row=i,column=ci,value=val); c.font=Font(name="Inter",size=10)
            c.fill=PatternFill("solid",fgColor=("F0F4FA" if i%2==0 else "FFFFFF"))
            if ci==2: c.number_format='#,##0'
            c.alignment=Alignment(horizontal="center" if ci>1 else "left",vertical="center")

    buf=io.BytesIO(); wb.save(buf); return buf.getvalue()

# ── Session defaults ──────────────────────────────────────────────────────────
for key,default in [("step",1),("brief",{}),("selected_client",None),("selected_brand",None),
                    ("combined_df",None),("generated_plan",None),("chat_messages",[]),
                    ("brief_summary",{}),("budget_split",{}),("channel_kpi_data",{}),
                    ("edit_mode",False),("edit_plan",{}),("edit_messages",[])]:
    if key not in st.session_state: st.session_state[key]=default

# ── Sidebar ───────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("""
    <div style='padding:16px 8px 8px 8px;'>
      <div style='font-family:Plus Jakarta Sans,sans-serif;font-size:1.3rem;font-weight:800;color:#fff;letter-spacing:-0.3px;'>🎯 Orchestration</div>
      <div style='font-size:0.72rem;color:rgba(255,255,255,0.5);margin-top:3px;letter-spacing:0.05em;text-transform:uppercase;'>Digital Planning Platform</div>
    </div>
    <hr style='margin:12px 0 16px 0;'>
    """, unsafe_allow_html=True)

    nav=st.radio("nav",[
        "📊  New Campaign Plan",
        "📁  Saved Plans",
        "📈  Data Explorer",
        "⚙️  Settings"
    ],label_visibility="collapsed")

    st.markdown("<br>",unsafe_allow_html=True)
    step=st.session_state.get("step",1)
    if nav=="📊  New Campaign Plan" and step>1:
        client_n=st.session_state.get("selected_client",{})
        brand_n=st.session_state.get("selected_brand",{})
        if client_n and brand_n:
            st.markdown(f"""
            <div style='background:rgba(255,255,255,0.1);border-radius:10px;padding:12px 14px;margin-bottom:8px;'>
              <div style='font-size:0.7rem;color:rgba(255,255,255,0.5);text-transform:uppercase;letter-spacing:0.06em;margin-bottom:4px;'>Active Session</div>
              <div style='font-size:0.85rem;font-weight:600;color:#fff;'>{client_n.get("client_name","")}</div>
              <div style='font-size:0.78rem;color:rgba(255,255,255,0.7);'>{brand_n.get("brand_name","")}</div>
              <div style='font-size:0.72rem;color:rgba(255,255,255,0.45);margin-top:4px;'>Step {step} of 4</div>
            </div>""", unsafe_allow_html=True)

    st.markdown("""
    <div style='position:absolute;bottom:20px;left:0;right:0;padding:0 16px;'>
      <div style='font-size:0.68rem;color:rgba(255,255,255,0.3);'>POC Version 1.0 · Phase 1</div>
    </div>""", unsafe_allow_html=True)

# ── Hero ──────────────────────────────────────────────────────────────────────
st.markdown("""
<div class="hero-header">
  <div class="hero-badge">Digital Planning Platform</div>
  <div class="hero-title">Orchestration-Digital</div>
  <div class="hero-sub">Plan · Analyse · Execute · Report — all in one place</div>
</div>""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# NEW CAMPAIGN PLAN
# ══════════════════════════════════════════════════════════════════════════════
if nav=="📊  New Campaign Plan":
    sb=get_supabase()
    # Scroll to top whenever this page loads
    components.html("<script>window.parent.document.querySelector('section.main').scrollTop=0;</script>",height=0)
    # If arriving at this tab while on step 4 (showing old plan) — reset to step 1
    prev_nav=st.session_state.get("prev_nav","")
    if prev_nav != "📊  New Campaign Plan" and st.session_state.get("step",1)==4:
        for k in ["step","brief","brief_form","selected_client","selected_brand","combined_df",
                  "loaded_benchmarks","generated_plan","chat_messages","brief_summary",
                  "auto_saved_version","budget_split","channel_kpi_data","budget_ranking",
                  "edit_mode","edit_plan","edit_messages"]:
            st.session_state[k] = 1 if k=="step" else ([] if k in ["chat_messages","edit_messages"] else None)
    st.session_state["prev_nav"]="📊  New Campaign Plan"
    steps=["1 · Client & Brand","2 · Historical Data","3 · Campaign Brief","4 · Media Plan"]
    cols=st.columns(4)
    for i,(col,label) in enumerate(zip(cols,steps),1):
        active=st.session_state["step"]==i; done=st.session_state["step"]>i
        bg="#1e3a5f" if active else ("#1a8a6e" if done else "#e8eaf0")
        fg="#ffffff" if (active or done) else "#8a93a8"
        col.markdown(f"""<div style='background:{bg};color:{fg};border-radius:10px;padding:12px 16px;text-align:center;font-weight:600;font-size:0.82rem;'>{"✓" if done else str(i)} · {label.split("·")[1].strip()}</div>""",unsafe_allow_html=True)
    st.markdown("<br>",unsafe_allow_html=True)

    # STEP 1
    if st.session_state["step"]==1:
        st.markdown('<div class="section-header">Select Client & Brand</div>',unsafe_allow_html=True)
        clients=get_clients_list(sb); client_names=[c["client_name"] for c in clients]
        col_a,col_b=st.columns(2)
        with col_a:
            st.markdown("**Client**")
            cc=st.selectbox("Client",["— Select —","➕ Add new client"]+client_names,label_visibility="collapsed")
            if cc=="➕ Add new client":
                nc=st.text_input("New client name",placeholder="e.g. Unilever Lanka")
                if st.button("Create Client") and nc:
                    r=create_client_record(sb,nc)
                    if r: st.success(f"✅ '{nc}' created!"); st.rerun()
            elif cc!="— Select —":
                st.session_state["selected_client"]=next((c for c in clients if c["client_name"]==cc),None)
        with col_b:
            if st.session_state.get("selected_client"):
                st.markdown("**Brand**")
                brands=get_brands_for_client(sb,st.session_state["selected_client"]["id"])
                brand_names=[b["brand_name"] for b in brands]
                bc=st.selectbox("Brand",["— Select —","➕ Add new brand"]+brand_names,label_visibility="collapsed")
                if bc=="➕ Add new brand":
                    nb=st.text_input("New brand name",placeholder="e.g. Sunlight")
                    if st.button("Create Brand") and nb:
                        r=create_brand_record(sb,st.session_state["selected_client"]["id"],nb)
                        if r: st.success(f"✅ '{nb}' created!"); st.rerun()
                elif bc!="— Select —":
                    st.session_state["selected_brand"]=next((b for b in brands if b["brand_name"]==bc),None)
            else: st.info("Select a client first.")
        if st.session_state.get("selected_client") and st.session_state.get("selected_brand"):
            st.markdown("---")
            st.markdown('<div class="green-btn">',unsafe_allow_html=True)
            if st.button("Continue to Historical Data →",use_container_width=True):
                st.session_state.update({"step":2,"chat_messages":[],"generated_plan":None,"budget_split":{},"channel_kpi_data":{},"brief_form":{},"brief_summary":{}})
                st.rerun()
            st.markdown('</div>',unsafe_allow_html=True)

    # STEP 2
    elif st.session_state["step"]==2:
        brand_id=st.session_state["selected_brand"]["id"]
        brand_name=st.session_state["selected_brand"]["brand_name"]
        client_name=st.session_state["selected_client"]["client_name"]
        st.markdown(f'<div class="section-header">Historical Data — {client_name} · {brand_name}</div>',unsafe_allow_html=True)

        st.markdown("""
        <div style='background:#f0f4fa;border-left:4px solid #2d5a9b;border-radius:0 10px 10px 0;padding:16px 20px;margin-bottom:20px;'>
          <div style='font-family:Plus Jakarta Sans,sans-serif;font-weight:700;color:#1e3a5f;font-size:0.92rem;margin-bottom:6px;'>📊 Why do we need your past campaign data?</div>
          <div style='font-size:0.85rem;color:#4a5168;line-height:1.7;'>
            Orchy uses your historical campaign data to make the plan <b>specific to your brand</b>, not generic.<br><br>
            From your past data, Orchy will:<br>
            &nbsp;&nbsp;• Calculate your <b>actual buying rates</b> — e.g. your real CPM on Facebook, not an industry estimate<br>
            &nbsp;&nbsp;• Set <b>realistic KPI targets</b> based on what your campaigns have actually achieved<br>
            &nbsp;&nbsp;• Identify your <b>most cost-efficient channels</b> to inform budget allocation<br><br>
            <b>What to upload:</b> Export your campaign performance reports from Meta Ads Manager, Google Ads, TikTok Ads Manager, or any other platform. CSV or Excel files work. The more historical data you provide, the more accurate the plan.
          </div>
        </div>
        """, unsafe_allow_html=True)

        saved_files=get_brand_data_files(sb,brand_id)
        if saved_files:
            total_rows=sum(f.get("row_count",0) for f in saved_files)
            st.success(f"✅ {len(saved_files)} saved file(s) for **{brand_name}** — auto-loaded.")
            c1,c2,c3=st.columns(3)
            with c1: st.markdown(f'<div class="metric-card"><div class="metric-value">{len(saved_files)}</div><div class="metric-label">Saved Files</div></div>',unsafe_allow_html=True)
            with c2: st.markdown(f'<div class="metric-card"><div class="metric-value">{total_rows:,}</div><div class="metric-label">Total Rows</div></div>',unsafe_allow_html=True)
            with c3: st.markdown(f'<div class="metric-card"><div class="metric-value">{brand_name}</div><div class="metric-label">Brand</div></div>',unsafe_allow_html=True)

            # FIX: load dfs OUTSIDE expander so combined_df is always set
            dfs=[]
            for f in saved_files:
                try: dfs.append(pd.read_json(io.StringIO(f["data_json"])))
                except: pass
            if dfs:
                st.session_state["combined_df"]=pd.concat(dfs,ignore_index=True)

            with st.expander("📁 Manage saved files"):
                for f in saved_files:
                    ca,cb,cc=st.columns([3,2,1])
                    ca.markdown(f"📄 **{f['file_name']}**")
                    cb.markdown(f"<span style='color:#8a93a8;font-size:0.82rem;'>{f.get('row_count',0):,} rows · {f['uploaded_at'][:10]}</span>",unsafe_allow_html=True)
                    if cc.button("🗑",key=f"del_{f['id']}"): delete_brand_data_file(sb,f["id"]); st.rerun()
        else:
            st.info(f"No historical data saved for **{brand_name}** yet.")

        st.markdown('<div class="section-header">Upload New Files</div>',unsafe_allow_html=True)
        uploaded_files=st.file_uploader("Drop files here",type=["csv","xlsx","xls"],accept_multiple_files=True)
        if uploaded_files:
            new_dfs=[(f.name,parse_uploaded_file(f)) for f in uploaded_files]
            new_dfs=[(n,d) for n,d in new_dfs if d is not None]
            if new_dfs:
                if st.button(f"💾 Save {len(new_dfs)} file(s) to {brand_name}",use_container_width=True):
                    for fname,df in new_dfs: save_brand_data(sb,brand_id,fname,df.to_json(),len(df))
                    st.success("✅ Saved!"); st.rerun()
                # FIX: always merge into combined_df immediately on upload
                new_combined=pd.concat([d for _,d in new_dfs],ignore_index=True)
                existing=st.session_state.get("combined_df")
                st.session_state["combined_df"]=pd.concat([existing,new_combined],ignore_index=True) if existing is not None else new_combined
                # FIX: clear loaded_benchmarks so benchmark table regenerates with new data
                st.session_state.pop("loaded_benchmarks", None)
        # ── Benchmark table from uploaded data ──────────────────────────────
        if st.session_state.get("combined_df") is not None:
            df_loaded = st.session_state["combined_df"]
            # Always recompute benchmarks fresh from current combined_df
            benchmarks_loaded = extract_channel_benchmarks(df_loaded)
            st.session_state["loaded_benchmarks"] = benchmarks_loaded

            st.markdown('<div class="section-header">📊 Performance Benchmarks from Your Data</div>', unsafe_allow_html=True)
            st.caption("Extracted from your uploaded data — used in planning to set buying rates and KPI targets. ⚠️ = Sri Lanka industry average (no historical data found).")

            obj_kpi_map = {"Awareness/Reach":"cpm","Video Views":"cpv","Traffic":"cpc","Leads/Conversions":"cpa"}
            bench_rows = []
            for ch, bench in benchmarks_loaded.items():
                if not bench or bench.get("rows",0)==0: continue
                is_global = bench.get("is_global_fallback", False)
                source = "⚠️ Industry Avg" if is_global else f"✅ Your Data ({bench.get('rows',0)} rows)"
                row = {"Channel": ch, "Source": source}
                for obj_label, metric in obj_kpi_map.items():
                    val = bench.get(metric)
                    if val:
                        row[obj_label] = f"LKR {val:,.0f}"
                    else:
                        ia_val = INDUSTRY_AVERAGES.get(ch,{}).get(metric)
                        row[obj_label] = f"LKR {ia_val:,.0f} ⚠️" if ia_val else "—"
                ctr = bench.get("ctr") or INDUSTRY_AVERAGES.get(ch,{}).get("ctr")
                row["CTR %"] = f"{ctr:.2f}%" if ctr else "—"
                roas = bench.get("roas") or INDUSTRY_AVERAGES.get(ch,{}).get("roas")
                row["ROAS"] = f"{roas:.1f}x" if roas else "—"
                bench_rows.append(row)

            if bench_rows:
                st.dataframe(pd.DataFrame(bench_rows), use_container_width=True, hide_index=True)
            else:
                st.info("Upload campaign data above to see performance benchmarks.")

        st.markdown("---")
        c1,c2=st.columns(2)
        with c1:
            if st.button("← Back",use_container_width=True): st.session_state["step"]=1; st.rerun()
        with c2:
            st.markdown('<div class="green-btn">',unsafe_allow_html=True)
            if st.button("Continue to Planning Chat →",use_container_width=True):
                if st.session_state.get("combined_df") is None: st.error("Please upload at least one data file.")
                else: st.session_state["step"]=3; st.rerun()
            st.markdown('</div>',unsafe_allow_html=True)

    # STEP 3 — CAMPAIGN BRIEF FORM
    elif st.session_state["step"]==3:
        client_name=st.session_state["selected_client"]["client_name"]
        brand_name=st.session_state["selected_brand"]["brand_name"]
        benchmarks=st.session_state.get("loaded_benchmarks") or extract_channel_benchmarks(st.session_state["combined_df"])

        # Scroll to top when brief form loads
        components.html("<script>window.parent.document.querySelector('section.main').scrollTop=0;</script>",height=0)
        st.markdown(f'<div class="section-header">Campaign Brief — {client_name} · {brand_name}</div>',unsafe_allow_html=True)

        ALL_CHANNELS=["Facebook","Instagram","YouTube","Google Search","Google Display","TikTok","LinkedIn","Programmatic Display"]
        OBJECTIVES=["Brand Awareness","Reach & Frequency","Video Views","Website Traffic","Lead Generation","E-commerce / Conversions","App Installs","Engagement"]
        # Asset types vary by channel — video platforms can't run static images
        ALL_ASSET_TYPES=["Static Images","Video — 6s","Video — 15s","Video — 30s","Video — 60s","Carousel","Stories / Reels","UGC Content","Influencer Content"]
        VIDEO_ONLY_CHANNELS=["youtube","tiktok"]  # no static images on these platforms
        def get_asset_types(ch):
            if any(v in ch.lower() for v in VIDEO_ONLY_CHANNELS):
                return [a for a in ALL_ASSET_TYPES if a!="Static Images"]
            return ALL_ASSET_TYPES
        ASSET_TYPES=ALL_ASSET_TYPES  # keep for reference

        # Only restore previous form if same brand/client — clear if new session
        prev_brand=st.session_state.get("brief_form",{}).get("brand_name","")
        prev=st.session_state.get("brief_form",{}) if prev_brand==brand_name else {}

        # ── Section 1: Campaign Basics ────────────────────────────────────────
        st.markdown('<div class="section-header">1 · Campaign Basics</div>',unsafe_allow_html=True)
        col1,col2,col3=st.columns(3)
        with col1:
            campaign_name=st.text_input("Campaign Name",value=prev.get("campaign_name",""),placeholder="e.g. Q3 2025 Brand Awareness")
        with col2:
            objective=st.selectbox("Campaign Objective",OBJECTIVES,index=OBJECTIVES.index(prev.get("objective","Brand Awareness")) if prev.get("objective") in OBJECTIVES else 0)
        with col3:
            market=st.text_input("Market / Region",value=prev.get("market","Sri Lanka"))

        col4,col5,col6=st.columns(3)
        with col4:
            budget_key="brief_budget_input"
            if budget_key not in st.session_state:
                st.session_state[budget_key]=int(prev.get("total_budget",0))
            total_budget=st.number_input(
                "Total Budget (LKR)",min_value=0,
                value=st.session_state[budget_key],
                step=50000,format="%d",key=budget_key)
            if total_budget>0:
                st.caption(f"💰 LKR {total_budget:,.0f}  ≈  USD {total_budget/320:,.0f}")
        with col5:
            start_date=st.date_input("Start Date",
                value=prev.get("start_date",date.today()),
                format="DD/MM/YYYY")
        with col6:
            end_date=st.date_input("End Date",
                value=prev.get("end_date",date.today()),
                format="DD/MM/YYYY")
            if end_date and start_date and end_date<=start_date:
                st.error("⚠️ End date must be after start date.")
        budget_type=st.radio("Budget Type",["Total campaign budget","Monthly budget"],horizontal=True,index=0 if prev.get("budget_type","Total")=="Total" else 1)

        # ── Section 2: Target Audience ────────────────────────────────────────
        st.markdown('<div class="section-header">2 · Target Audience</div>',unsafe_allow_html=True)
        col7,col8,col9=st.columns(3)
        with col7:
            age_col1,age_col2=st.columns(2)
            age_min=age_col1.number_input("Age Min",min_value=13,max_value=65,value=int(prev.get("age_min",18)))
            age_max=age_col2.number_input("Age Max",min_value=13,max_value=65,value=int(prev.get("age_max",45)))
        with col8:
            gender=st.selectbox("Gender",["All","Male","Female"],index=["All","Male","Female"].index(prev.get("gender","All")))
        with col9:
            locations=st.text_input("Locations",value=prev.get("locations","Sri Lanka — Island-wide"),placeholder="e.g. Colombo, Kandy, Galle")

        col10,col11=st.columns(2)
        with col10:
            interests=st.text_area("Interests & Behaviours",value=prev.get("interests",""),placeholder="e.g. Finance, Tech, Online Shopping",height=80)
        with col11:
            custom_audiences=st.text_area("Custom Audiences",value=prev.get("custom_audiences",""),placeholder="e.g. Website visitors, CRM list upload",height=40)
            lookalike=st.text_input("Lookalike Audiences",value=prev.get("lookalike",""),placeholder="e.g. 1% lookalike of past purchasers")

        # ── Section 3: Channel Selection ──────────────────────────────────────
        st.markdown('<div class="section-header">3 · Channel Selection</div>',unsafe_allow_html=True)
        st.caption("Select channels to include. For each selected channel, fill in the audience size from the platform estimator and the primary KPI.")

        prev_channels=prev.get("channels",[])
        selected_channels=[]
        channel_audience_sizes=prev.get("audience_sizes",{})
        channel_kpis_form=prev.get("channel_kpis",{})
        channel_dates_form=prev.get("channel_dates",{})
        channel_instructions=prev.get("channel_instructions",{})

        KPI_OPTIONS=["CPM (Awareness/Reach)","CPC (Traffic/Clicks)","CPV (Video Views)","CPL (Leads)","CPA (Conversions)","ROAS (E-commerce)"]

        for ch in ALL_CHANNELS:
            ch_selected=st.checkbox(f"**{ch}**",value=ch in prev_channels,key=f"ch_{ch}")
            if ch_selected:
                selected_channels.append(ch)
                with st.container():
                    bench=benchmarks.get(ch,{})
                    industry=INDUSTRY_AVERAGES.get(ch,{})
                    cpm_hint=bench.get("cpm") or industry.get("cpm","—")
                    cpc_hint=bench.get("cpc") or industry.get("cpc","—")
                    hint_txt=f"Historical: CPM={cpm_hint}, CPC={cpc_hint}" if cpm_hint!="—" else "No historical data — industry averages will be used"
                    st.caption(f"📊 {hint_txt}")
                    cc1,cc2,cc3,cc4=st.columns(4)
                    with cc1:
                        aud_val=channel_audience_sizes.get(ch,0)
                        aud_input=st.number_input(
                            f"Targetable Audience",min_value=0,value=int(aud_val),
                            step=1000,format="%d",key=f"aud_{ch}",
                            help="Get this from Meta Audience Insights / Google Reach Planner / TikTok Audience Estimator"
                        )
                        channel_audience_sizes[ch]=aud_input
                        if aud_input>0:
                            st.caption(f"{aud_input:,.0f} people")
                    with cc2:
                        prev_kpi=channel_kpis_form.get(ch,"CPM (Awareness/Reach)")
                        kpi_idx=KPI_OPTIONS.index(prev_kpi) if prev_kpi in KPI_OPTIONS else 0
                        channel_kpis_form[ch]=st.selectbox(f"Primary KPI",KPI_OPTIONS,index=kpi_idx,key=f"kpi_{ch}")
                    with cc3:
                        channel_dates_form[ch]=st.text_input(f"Flight Dates",value=channel_dates_form.get(ch,"Same as campaign"),key=f"dates_{ch}")
                    with cc4:
                        channel_instructions[ch]=st.text_input(f"Special Instructions",value=channel_instructions.get(ch,""),key=f"inst_{ch}",placeholder="e.g. Mobile only")
                    st.markdown("<div style='height:4px'></div>",unsafe_allow_html=True)

        # ── Section 4: Creative Assets ─────────────────────────────────────────
        if selected_channels:
            st.markdown('<div class="section-header">4 · Creative Assets</div>',unsafe_allow_html=True)
            st.caption("Select available creative assets per channel. Optionally specify budget split by asset type.")
            prev_assets=prev.get("creative_assets",{})
            prev_asset_budgets=prev.get("asset_budgets",{})
            creative_assets={}
            asset_budgets={}
            split_by_asset=st.checkbox("Split budget by creative asset type",value=prev.get("split_by_asset",False))

            for ch in selected_channels:
                ch_asset_types=get_asset_types(ch)
                st.markdown(f"**{ch}**")
                if any(v in ch.lower() for v in VIDEO_ONLY_CHANNELS):
                    st.caption("ℹ️ Static Images not available on this platform")
                cols=st.columns(len(ch_asset_types))
                selected_assets=[]
                for i,asset in enumerate(ch_asset_types):
                    if cols[i].checkbox(asset.replace(" — "," "),value=asset in prev_assets.get(ch,[]),key=f"asset_{ch}_{i}",label_visibility="visible"):
                        selected_assets.append(asset)
                creative_assets[ch]=selected_assets

                if split_by_asset and selected_assets:
                    st.caption(f"Budget split for {ch} (LKR) — must add up to channel total:")
                    ab_cols=st.columns(min(len(selected_assets),4))
                    ch_budgets={}
                    for i,asset in enumerate(selected_assets):
                        col_idx=i%4
                        ab_val=int(prev_asset_budgets.get(ch,{}).get(asset,0))
                        ab_input=ab_cols[col_idx].number_input(
                            asset,min_value=0,value=ab_val,
                            step=10000,format="%d",key=f"ab_{ch}_{i}",label_visibility="visible"
                        )
                        if ab_input>0:
                            ab_cols[col_idx].caption(f"LKR {ab_input:,.0f}")
                        ch_budgets[asset]=ab_input
                    asset_budgets[ch]=ch_budgets
                st.markdown("<div style='height:4px'></div>",unsafe_allow_html=True)

        # ── Section 5: Planning Instructions ──────────────────────────────────
        st.markdown('<div class="section-header">5 · Additional Planning Instructions</div>',unsafe_allow_html=True)
        planning_instructions=st.text_area(
            "Instructions for the AI planner",
            value=prev.get("planning_instructions",""),
            placeholder="e.g. Prioritise reach over frequency. Avoid competitor brand keywords. Focus on mobile placements. Weight Facebook higher than Instagram.",
            height=100,
            label_visibility="collapsed"
        )

        # ── Generate ──────────────────────────────────────────────────────────
        st.markdown("---")
        col_back,col_gen=st.columns([1,3])
        with col_back:
            if st.button("← Back",use_container_width=True):
                st.session_state["step"]=2; st.rerun()
        with col_gen:
            st.markdown('<div class="green-btn">',unsafe_allow_html=True)
            if st.button("🚀 Generate Media Plan",use_container_width=True):
                # Validate
                if not campaign_name:
                    st.error("Please enter a Campaign Name.")
                elif end_date<=start_date:
                    st.error("End date must be after start date.")
                elif not selected_channels:
                    st.error("Please select at least one channel.")
                elif any(channel_audience_sizes.get(ch,0)==0 for ch in selected_channels):
                    st.warning("⚠️ Some channels have no audience size entered — budget optimisation will use estimates. Continue anyway?")
                else:
                    # Build brief dict
                    brief={
                        "campaign_name":campaign_name,
                        "objective":objective,
                        "total_budget":total_budget,
                        "start_date":str(start_date),
                        "end_date":str(end_date),
                        "market":market,
                        "budget_type":budget_type,
                        "age_min":age_min,
                        "age_max":age_max,
                        "gender":gender,
                        "locations":locations,
                        "interests":interests,
                        "custom_audiences":custom_audiences,
                        "lookalike":lookalike,
                        "channels":selected_channels,
                        "audience_sizes":{ch:int(v) for ch,v in channel_audience_sizes.items() if ch in selected_channels},
                        "channel_kpis":{ch:v for ch,v in channel_kpis_form.items() if ch in selected_channels},
                        "channel_dates":{ch:v for ch,v in channel_dates_form.items() if ch in selected_channels},
                        "channel_instructions":{ch:v for ch,v in channel_instructions.items() if ch in selected_channels},
                        "creative_assets":{ch:v for ch,v in creative_assets.items() if ch in selected_channels},
                        "asset_budgets":asset_budgets if split_by_asset else {},
                        "split_by_asset":split_by_asset,
                        "planning_instructions":planning_instructions,
                    }
                    st.session_state["brief_form"]=brief
                    st.session_state["brief_summary"]=brief

                    with st.spinner("🤖 Optimising budget split and generating your media plan…"):
                        try:
                            data_summary=summarise_dataframe(st.session_state["combined_df"])
                            # Merge benchmarks: client manual > uploaded data > industry average
                            client_id=st.session_state["selected_client"]["id"]
                            client_bm=get_client_benchmarks_as_dict(sb,client_id)
                            merged_benchmarks={**benchmarks}
                            for plat,bm_data in client_bm.items():
                                if any(bm_data.get(m) for m in ["cpm","cpc","cpv","cpa"]):
                                    merged_benchmarks[plat]={**merged_benchmarks.get(plat,{}),**bm_data}
                            benchmarks=merged_benchmarks
                            # Extract clean KPI type strings
                            ch_kpi_types={ch:v.split(" ")[0] for ch,v in channel_kpis_form.items() if ch in selected_channels}
                            budget_split=calculate_budget_split(
                                selected_channels,total_budget,objective,
                                {ch:int(channel_audience_sizes.get(ch,500000)) for ch in selected_channels},
                                benchmarks,ch_kpi_types
                            )
                            st.session_state["budget_split"]=budget_split

                            # Per-asset budget placements
                            channel_kpi_data={}
                            for ch,b in budget_split.items():
                                kpi=calculate_channel_kpis(ch,b,benchmarks,objective)
                                # Attach per-asset placements if split by asset
                                if split_by_asset and ch in asset_budgets and asset_budgets[ch]:
                                    enriched=[]
                                    for asset,ab in asset_budgets[ch].items():
                                        if ab>0:
                                            pl_kpi=calculate_channel_kpis(ch,ab,benchmarks,objective)
                                            enriched.append({
                                                "placement":asset,"budget":ab,
                                                "kpi_type":kpi["kpi_type"],
                                                "buying_rate":pl_kpi["buying_rate"],
                                                "target_kpi":pl_kpi["target_kpi"],
                                                "assets":asset,"objective":objective,
                                            })
                                    if enriched: kpi["placements"]=enriched
                                channel_kpi_data[ch]=kpi
                            st.session_state["channel_kpi_data"]=channel_kpi_data

                            data_gaps=get_data_gaps(selected_channels,benchmarks)
                            plan_text=generate_media_plan(brief,client_name,brand_name,data_summary,budget_split,channel_kpi_data,data_gaps)
                            st.session_state["generated_plan"]=plan_text

                            pv_auto=get_plan_version(sb,campaign_name,st.session_state["selected_brand"]["id"])
                            save_plan(sb,{
                                "brand_id":st.session_state["selected_brand"]["id"],
                                "client_name":client_name,"brand_name":brand_name,
                                "campaign_name":campaign_name,"objective":objective,
                                "total_budget":float(total_budget),
                                "start_date":str(start_date),"end_date":str(end_date),
                                "channels":json.dumps(selected_channels),
                                "market":market,"kpi_focus":"",
                                "plan_text":plan_text,"plan_version":pv_auto,
                                "budget_split":json.dumps(budget_split),
                                "channel_kpi_data":json.dumps(channel_kpi_data),
                                "planning_conversation":json.dumps(brief),
                                "created_at":datetime.utcnow().isoformat()
                            })
                            st.session_state["auto_saved_version"]=pv_auto
                            st.session_state["step"]=4
                            st.rerun()
                        except Exception as e:
                            st.error(f"Error generating plan: {e}")
            st.markdown('</div>',unsafe_allow_html=True)

    # STEP 4
    elif st.session_state["step"]==4:
        components.html("<script>window.parent.document.querySelector('section.main').scrollTop=0;</script>",height=0)
        client_name=st.session_state["selected_client"]["client_name"]
        brand_name=st.session_state["selected_brand"]["brand_name"]
        plan_text=st.session_state["generated_plan"]
        brief_summary=st.session_state.get("brief_summary",{})
        budget_split=st.session_state.get("budget_split",{})
        channel_kpi_data=st.session_state.get("channel_kpi_data",{})
        total_budget=float(brief_summary.get("total_budget",0))

        st.markdown(f'<div class="section-header">{client_name} · {brand_name} · {brief_summary.get("campaign_name","Media Plan")}</div>',unsafe_allow_html=True)

        try:
            s=datetime.strptime(brief_summary.get("start_date",""),"%Y-%m-%d")
            e=datetime.strptime(brief_summary.get("end_date",""),"%Y-%m-%d"); days=(e-s).days
        except: days="—"

        c1,c2,c3,c4=st.columns(4)
        with c1: st.markdown(f'<div class="metric-card"><div class="metric-value">{format_lkr(total_budget)}</div><div class="metric-label">Total Budget</div></div>',unsafe_allow_html=True)
        with c2: st.markdown(f'<div class="metric-card"><div class="metric-value">{days} days</div><div class="metric-label">Duration</div></div>',unsafe_allow_html=True)
        with c3: st.markdown(f'<div class="metric-card"><div class="metric-value">{len(budget_split)}</div><div class="metric-label">Channels</div></div>',unsafe_allow_html=True)
        with c4:
            obj=brief_summary.get("objective","—")
            st.markdown(f'<div class="metric-card"><div class="metric-value">{obj.split()[0] if obj else "—"}</div><div class="metric-label">Objective</div></div>',unsafe_allow_html=True)

        # Data gaps warning
        benchmarks=extract_channel_benchmarks(st.session_state["combined_df"]) if st.session_state.get("combined_df") is not None else {}
        gaps=get_data_gaps(list(budget_split.keys()),benchmarks)
        if gaps:
            st.warning("⚠️ **Industry averages used for the following channels** (no historical data found):\n\n"+"\n\n".join(gaps))

        if budget_split:
            st.markdown('<div class="section-header">Budget Split & KPI Targets</div>',unsafe_allow_html=True)

            # Show channel ranking by CPR if available
            ranking=st.session_state.get("budget_ranking",[])
            if ranking:
                st.caption("📊 Channels ranked by Cost Per Result (lowest CPR = highest priority = more budget)")
                rank_rows=[]
                for i,r in enumerate(ranking,1):
                    cap_note=""
                    if r["allocated"]>=r["budget_80pct"]*0.99:
                        cap_note="🔒 80% reach cap"
                    rank_rows.append({
                        "Rank":f"#{i}",
                        "Channel":r["channel"],
                        "CPR (LKR)":f"LKR {r['cpr']:,.2f}",
                        "Targetable Audience":f"{r['audience']:,}",
                        "Budget for 80% Reach":f"LKR {r['budget_80pct']:,.0f}",
                        "Allocated Budget":f"LKR {r['allocated']:,.0f}",
                        "Note":cap_note,
                    })
                st.dataframe(pd.DataFrame(rank_rows),use_container_width=True,hide_index=True)
                st.markdown("<br>",unsafe_allow_html=True)

            rows=[]
            for ch,budget in budget_split.items():
                kpi=channel_kpi_data.get(ch,{})
                rows.append({"Channel":ch,"Budget (LKR)":f"LKR {budget:,.0f}",
                             "% of Total":f"{budget/total_budget*100:.1f}%" if total_budget>0 else "—",
                             "KPI Type":kpi.get("kpi_type","—"),"Buying Rate":kpi.get("buying_rate","—"),
                             "Target KPI":kpi.get("target_kpi","—"),"Data Source":kpi.get("data_source","—")})
            st.dataframe(pd.DataFrame(rows),use_container_width=True,hide_index=True)

        # ── Action buttons at the TOP ────────────────────────────────────────
        pv=st.session_state.get("auto_saved_version","V1.0")
        excel_bytes=build_excel(brief_summary,plan_text,client_name,brand_name,budget_split,channel_kpi_data,pv)
        ca,cb,cc=st.columns(3)
        with ca:
            st.download_button("📥 Download TXT",data=plan_text,
                file_name=f"{brief_summary.get('campaign_name','plan').replace(' ','_')}_plan.txt",
                mime="text/plain",use_container_width=True)
        with cb:
            st.download_button("📊 Download Excel",data=excel_bytes,
                file_name=f"{brief_summary.get('campaign_name','plan').replace(' ','_')}_MediaPlan.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",use_container_width=True)
        with cc:
            st.markdown('<div class="green-btn">',unsafe_allow_html=True)
            if st.button("➕ New Plan",use_container_width=True):
                for k in ["step","brief","brief_form","selected_client","selected_brand","combined_df",
                          "loaded_benchmarks","generated_plan","chat_messages","brief_summary",
                          "auto_saved_version","budget_split","channel_kpi_data","budget_ranking",
                          "edit_mode","edit_plan","edit_messages","brief_budget_input"]:
                    st.session_state[k] = 1 if k=="step" else ([] if k in ["chat_messages","edit_messages"] else None)
                st.rerun()
            st.markdown('</div>',unsafe_allow_html=True)
        st.success(f"✅ Plan auto-saved to library as **{pv}**")

        st.markdown("---")

        # ── Plan rendered as structured sections with tables ──────────────────
        st.markdown('<div class="section-header">Media Plan</div>',unsafe_allow_html=True)

        # Split plan into sections and render with st.markdown (handles tables natively)
        sections = plan_text.split("## ")
        if len(sections) <= 1:
            # Fallback if no ## headers — render in expander to keep it compact
            with st.expander("📄 View Full Plan", expanded=True):
                st.markdown(plan_text)
        else:
            for section in sections:
                if not section.strip(): continue
                lines = section.strip().split("\n",1)
                title = lines[0].strip()
                body = lines[1].strip() if len(lines)>1 else ""
                # Show exec summary expanded, rest collapsed
                expanded = any(k in title.lower() for k in ["executive","summary","channel plan","kpi"])
                with st.expander(f"**{title}**", expanded=expanded):
                    st.markdown(body)

# ══════════════════════════════════════════════════════════════════════════════
# SAVED PLANS
# ══════════════════════════════════════════════════════════════════════════════
elif nav=="📁  Saved Plans":
    st.session_state["prev_nav"]="📁  Saved Plans"
    sb=get_supabase()

    # Edit mode
    if st.session_state.get("edit_mode") and st.session_state.get("edit_plan"):
        ep=st.session_state["edit_plan"]
        client_name=ep.get("client_name","")
        brand_name=ep.get("brand_name","")
        original_plan=ep.get("plan_text","")
        budget_split=json.loads(ep.get("budget_split","{}"))
        channel_kpi_data=json.loads(ep.get("channel_kpi_data","{}"))
        original_conversation=json.loads(ep.get("planning_conversation","[]"))
        original_conv_text="\n".join([
            f"{'PLANNER' if m['role']=='user' else 'ORCHY'}: {m['content']}"
            for m in original_conversation if m.get("role") in ["user","agent"]
        ])

        st.markdown(f'<div class="section-header">✏️ Editing — {client_name} · {brand_name} · {ep.get("campaign_name","")} {ep.get("plan_version","")}</div>',unsafe_allow_html=True)
        st.caption("Tell Orchy what you'd like to change. Be specific about budgets, channels, or strategy adjustments.")

        if not st.session_state["edit_messages"]:
            with st.spinner("Orchy is reviewing the plan…"):
                first=get_agent_response(
                    [{"role":"user","content":"I need to edit this campaign plan. Please review it and ask me what I'd like to change."}],
                    client_name,brand_name,"",mode="editing",
                    existing_plan=original_plan,
                    original_conversation=original_conv_text
                )
            st.session_state["edit_messages"]=[
                {"role":"user","content":"I need to edit this campaign plan. Please review it and ask me what I'd like to change."},
                {"role":"agent","content":first}
            ]
            st.rerun()

        # Render edit chat using native components
        for msg in st.session_state["edit_messages"]:
            role_display="assistant" if msg["role"]=="agent" else "user"
            avatar="🤖" if msg["role"]=="agent" else "🧑‍💼"
            with st.chat_message(role_display,avatar=avatar):
                st.markdown(msg["content"].replace("[EDIT_COMPLETE]","").strip())

        last_agent=next((m["content"] for m in reversed(st.session_state["edit_messages"]) if m["role"]=="agent"),"")
        edit_complete="[EDIT_COMPLETE]" in last_agent
        num_edit_msgs=len([m for m in st.session_state["edit_messages"] if m["role"]=="user"])

        # Native chat input for edit mode
        if edit_input := st.chat_input("Tell Orchy what to change… (Enter to send, Shift+Enter for new line)"):
            st.session_state["edit_messages"].append({"role":"user","content":edit_input.strip()})
            with st.spinner("Orchy is thinking…"):
                reply=get_agent_response(
                    st.session_state["edit_messages"],client_name,brand_name,"",
                    mode="editing",existing_plan=original_plan,
                    original_conversation=original_conv_text
                )
            st.session_state["edit_messages"].append({"role":"agent","content":reply})
            st.rerun()

        # Apply / Cancel below input
        if num_edit_msgs>=1:
            st.markdown("---")
            if edit_complete:
                st.success("✅ Orchy has noted all the changes. Ready to generate the updated plan!")
            else:
                st.info("💡 Keep adding instructions, or click below when ready to regenerate.")
            st.markdown('<div class="green-btn">',unsafe_allow_html=True)
            if st.button("🚀 Apply Edits & Generate New Version",use_container_width=True):
                with st.spinner("🤖 Applying edits and generating updated plan…"):
                    try:
                        new_plan=apply_plan_edits(st.session_state["edit_messages"],original_plan,client_name,brand_name,budget_split,channel_kpi_data,original_conv_text)
                        new_version=get_plan_version(sb,ep.get("campaign_name",""),ep.get("brand_id",""))
                        save_plan(sb,{"brand_id":ep.get("brand_id"),"client_name":client_name,"brand_name":brand_name,
                            "campaign_name":ep.get("campaign_name",""),"objective":ep.get("objective",""),
                            "total_budget":ep.get("total_budget",0),"start_date":ep.get("start_date",""),
                            "end_date":ep.get("end_date",""),"channels":ep.get("channels","[]"),
                            "market":ep.get("market","Sri Lanka"),"kpi_focus":"","plan_text":new_plan,
                            "plan_version":new_version,"budget_split":ep.get("budget_split","{}"),
                            "channel_kpi_data":ep.get("channel_kpi_data","{}"),
                            "created_at":datetime.utcnow().isoformat()})
                        st.success(f"✅ New version {new_version} saved to library!")
                        st.markdown('<div class="plan-card">'+new_plan.replace("\n","<br>")+"</div>",unsafe_allow_html=True)
                        st.session_state.update({"edit_mode":False,"edit_plan":{},"edit_messages":[]})
                    except Exception as e: st.error(f"Error: {e}")
            st.markdown('</div>',unsafe_allow_html=True)



    else:
        st.markdown('<div class="section-header">Saved Campaign Plans</div>',unsafe_allow_html=True)
        try:
            plans=load_saved_plans(sb)
            if not plans:
                st.info("No saved plans yet. Create your first plan to see it here.")
            else:
                # Group by campaign
                for plan in plans:
                    budget=float(plan.get("total_budget",0) or 0)
                    version=plan.get("plan_version","V1.0")
                    label=f"📋 {plan.get('client_name','—')} · {plan.get('brand_name','—')} · {plan.get('campaign_name','Unnamed')} [{version}] — {format_lkr(budget)}"
                    with st.expander(label):
                        c1,c2,c3,c4=st.columns(4)
                        c1.metric("Objective",plan.get("objective","—"))
                        c2.metric("Budget",format_lkr(budget))
                        c3.metric("Version",version)
                        c4.metric("Market",plan.get("market","—"))
                        st.markdown(f"**Dates:** {plan.get('start_date','')} → {plan.get('end_date','')}")

                        # KPI table if available
                        if plan.get("channel_kpi_data"):
                            try:
                                kpi_data=json.loads(plan["channel_kpi_data"])
                                budget_data=json.loads(plan.get("budget_split","{}"))
                                if kpi_data:
                                    rows=[{"Channel":ch,"Budget":format_lkr(budget_data.get(ch,0)),
                                           "KPI Type":v.get("kpi_type","—"),"Buying Rate":v.get("buying_rate","—"),
                                           "Target KPI":v.get("target_kpi","—")} for ch,v in kpi_data.items()]
                                    st.dataframe(pd.DataFrame(rows),use_container_width=True,hide_index=True)
                            except: pass

                        st.markdown("---")
                        st.markdown(plan.get("plan_text",""))
                        st.markdown("---")

                        col_edit2,col_del=st.columns(2)
                        with col_edit2:
                            if st.button(f"✏️ Edit this Plan",key=f"edit_{plan['id']}",use_container_width=True):
                                st.session_state.update({"edit_mode":True,"edit_plan":plan,"edit_messages":[],"edit_original_conversation":json.loads(plan.get("planning_conversation","[]"))})
                                st.rerun()
                        with col_del:
                            if st.button(f"🗑️ Delete",key=f"del_plan_{plan['id']}",use_container_width=True):
                                st.session_state[f"confirm_del_{plan['id']}"]=True
                                st.rerun()
                        if st.session_state.get(f"confirm_del_{plan['id']}"):
                            st.warning(f"Are you sure you want to delete **{plan.get('campaign_name','')} {plan.get('plan_version','')}**? This cannot be undone.")
                            cy,cn=st.columns(2)
                            if cy.button("Yes, delete it",key=f"yes_del_{plan['id']}"):
                                delete_plan(sb,plan["id"])
                                st.session_state.pop(f"confirm_del_{plan['id']}",None)
                                st.success("Plan deleted.")
                                st.rerun()
                            if cn.button("Cancel",key=f"no_del_{plan['id']}"):
                                st.session_state.pop(f"confirm_del_{plan['id']}",None)
                                st.rerun()
                        try:
                            bs=json.loads(plan.get("budget_split","{}"))
                            ck=json.loads(plan.get("channel_kpi_data","{}"))
                            bs_obj={"campaign_name":plan.get("campaign_name",""),"objective":plan.get("objective",""),
                                    "start_date":plan.get("start_date",""),"end_date":plan.get("end_date",""),"audience":""}
                            excel=build_excel(bs_obj,plan.get("plan_text",""),plan.get("client_name",""),plan.get("brand_name",""),bs,ck,version)
                            st.download_button("📊 Download Excel",data=excel,
                                file_name=f"{plan.get('campaign_name','plan').replace(' ','_')}_{version}_MediaPlan.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True,key=f"dl_{plan['id']}")
                        except: pass
        except Exception as e: st.error(f"Could not load plans: {e}")

# ══════════════════════════════════════════════════════════════════════════════
# DATA EXPLORER
# ══════════════════════════════════════════════════════════════════════════════
elif nav=="📈  Data Explorer":
    st.markdown('<div class="section-header">Data Explorer</div>',unsafe_allow_html=True)
    if st.session_state.get("combined_df") is None:
        st.info("Complete Step 2 in New Campaign Plan to load data here.")
    else:
        df=st.session_state["combined_df"]
        st.caption(f"{len(df):,} rows · {len(df.columns)} columns")
        st.dataframe(df,use_container_width=True)
        numeric_cols=df.select_dtypes(include='number').columns.tolist()
        if len(numeric_cols)>=2:
            st.markdown('<div class="section-header">Quick Chart</div>',unsafe_allow_html=True)
            c1,c2=st.columns(2)
            x_col=c1.selectbox("X axis",df.columns.tolist())
            y_col=c2.selectbox("Y axis",numeric_cols)
            chart_type=st.radio("Chart type",["Line","Bar","Area"],horizontal=True)
            chart_df=df[[x_col,y_col]].dropna().set_index(x_col)
            if chart_type=="Line": st.line_chart(chart_df)
            elif chart_type=="Bar": st.bar_chart(chart_df)
            else: st.area_chart(chart_df)

# ══════════════════════════════════════════════════════════════════════════════
# SETTINGS
# ══════════════════════════════════════════════════════════════════════════════
elif nav=="⚙️  Settings":
    sb=get_supabase()
    st.markdown('<div class="section-header">Platform Settings</div>',unsafe_allow_html=True)
    st.caption("Configure platform efficiency benchmarks used to optimise budget allocation. Audience sizes are captured per campaign during the Orchy planning chat.")

    platforms=[
        {"name":"Facebook","default_freq_awareness":3,"default_freq_conversion":7,"reach_curve_inflection":0.4},
        {"name":"Instagram","default_freq_awareness":3,"default_freq_conversion":6,"reach_curve_inflection":0.35},
        {"name":"YouTube","default_freq_awareness":4,"default_freq_conversion":5,"reach_curve_inflection":0.45},
        {"name":"Google Search","default_freq_awareness":1,"default_freq_conversion":3,"reach_curve_inflection":0.6},
        {"name":"Google Display","default_freq_awareness":5,"default_freq_conversion":8,"reach_curve_inflection":0.3},
        {"name":"TikTok","default_freq_awareness":3,"default_freq_conversion":5,"reach_curve_inflection":0.38},
        {"name":"LinkedIn","default_freq_awareness":4,"default_freq_conversion":6,"reach_curve_inflection":0.5},
        {"name":"Programmatic","default_freq_awareness":6,"default_freq_conversion":9,"reach_curve_inflection":0.3},
    ]
    existing_settings=get_platform_settings(sb)
    existing_dict={p["platform_name"]:p for p in existing_settings}

    st.markdown('<div class="section-header">Frequency Caps & Reach Curve Parameters</div>',unsafe_allow_html=True)
    st.caption("Frequency cap = max times a user sees your ad per campaign. Reach curve inflection = point (as % of targetable audience) where diminishing returns begin.")

    for p in platforms:
        pname=p["name"]; saved=existing_dict.get(pname,{})
        with st.expander(f"⚙️ {pname}"):
            col1,col2,col3=st.columns(3)
            fa=col1.number_input("Freq Cap — Awareness",min_value=1,max_value=20,value=int(saved.get("freq_cap_awareness",p["default_freq_awareness"])),key=f"fa_{pname}")
            fc=col2.number_input("Freq Cap — Conversion",min_value=1,max_value=20,value=int(saved.get("freq_cap_conversion",p["default_freq_conversion"])),key=f"fc_{pname}")
            ri=col3.number_input("Reach Curve Inflection (%)",min_value=0.1,max_value=1.0,step=0.05,value=float(saved.get("reach_curve_inflection",p["reach_curve_inflection"])),key=f"ri_{pname}")
            notes=st.text_input("Notes",value=saved.get("notes",""),key=f"n_{pname}",placeholder="e.g. Strong for video in LK, peak hours 7–10pm")
            if st.button(f"Save {pname}",key=f"save_{pname}"):
                save_platform_setting(sb,pname,{"freq_cap_awareness":fa,"freq_cap_conversion":fc,"reach_curve_inflection":ri,"notes":notes})
                st.success(f"✅ {pname} settings saved!")

    # ── Client Historical Benchmarks ─────────────────────────────────────────
    st.markdown('<div class="section-header">Client Historical Benchmarks</div>',unsafe_allow_html=True)
    st.caption("Enter historical buying rates per client. These take priority over uploaded data and industry averages when planning for that client.")

    clients_list=get_clients_list(sb)
    if not clients_list:
        st.info("No clients found. Create a client first in New Campaign Plan.")
    else:
        client_names_list=[c["client_name"] for c in clients_list]
        sel_client_name=st.selectbox("Select Client",client_names_list,key="settings_client")
        sel_client=next((c for c in clients_list if c["client_name"]==sel_client_name),None)

        if sel_client:
            existing_benchmarks=get_client_benchmarks_as_dict(sb,sel_client["id"])
            PLATFORMS=["Facebook","Instagram","YouTube","Google Search","Google Display","TikTok","LinkedIn","Programmatic Display"]
            METRICS=["cpm","cpc","cpv","cpa","ctr","roas"]
            METRIC_LABELS={"cpm":"CPM (LKR)","cpc":"CPC (LKR)","cpv":"CPV (LKR)","cpa":"CPA (LKR)","ctr":"CTR (%)","roas":"ROAS"}

            st.markdown(f"**Benchmarks for {sel_client_name}**")
            st.caption("Leave at 0 to use uploaded data or industry averages. Values entered here will be used for all plans for this client.")

            for platform in PLATFORMS:
                existing=existing_benchmarks.get(platform,{})
                industry=INDUSTRY_AVERAGES.get(platform,{})
                with st.expander(f"📊 {platform}" + (" ✅ Custom data saved" if existing.get("cpm") or existing.get("cpc") else " — using defaults")):
                    # Show current values vs industry averages
                    st.markdown("<div style='font-size:0.8rem;color:#8a93a8;margin-bottom:8px;'>Industry average shown as placeholder. Enter your client's actual rates.</div>",unsafe_allow_html=True)
                    cols=st.columns(3)
                    vals={}
                    for i,metric in enumerate(METRICS):
                        col=cols[i%3]
                        ia_val=industry.get(metric,0) or 0
                        saved_val=existing.get(metric) or 0
                        display_val=float(saved_val) if saved_val else 0.0
                        vals[metric]=col.number_input(
                            METRIC_LABELS[metric],
                            min_value=0.0,value=display_val,
                            step=1.0 if metric not in ["ctr","roas"] else 0.01,
                            format="%.2f" if metric in ["ctr","roas","cpv"] else "%.0f",
                            key=f"bm_{sel_client['id']}_{platform}_{metric}",
                            help=f"Industry avg: {ia_val}"
                        )
                    sample_size=st.number_input("Sample size (number of campaigns this is based on)",
                        min_value=0,value=int(existing.get("sample_size",0) or 0),
                        key=f"bm_ss_{sel_client['id']}_{platform}")
                    if st.button(f"💾 Save {platform}",key=f"save_bm_{sel_client['id']}_{platform}"):
                        save_client_benchmark(sb,sel_client["id"],platform,{
                            "cpm":vals["cpm"] or None,"cpc":vals["cpc"] or None,
                            "cpv":vals["cpv"] or None,"cpa":vals["cpa"] or None,
                            "ctr":vals["ctr"] or None,"roas":vals["roas"] or None,
                            "sample_size":sample_size
                        })
                        st.success(f"✅ {platform} benchmarks saved for {sel_client_name}!")
                        st.rerun()

    st.markdown('<div class="section-header">Sri Lanka Industry Averages (Fallback)</div>',unsafe_allow_html=True)
    st.caption("These are used when no historical data exists for a channel. Clearly marked in plan outputs.")
    rows=[{"Platform":k,"CPM (LKR)":v.get("cpm","—"),"CPC (LKR)":v.get("cpc","—"),"CPV (LKR)":v.get("cpv","—"),"CTR (%)":v.get("ctr","—"),"ROAS":v.get("roas","—")} for k,v in INDUSTRY_AVERAGES.items()]
    st.dataframe(pd.DataFrame(rows),use_container_width=True,hide_index=True)

    st.markdown('<div class="section-header">About Budget Optimisation</div>',unsafe_allow_html=True)
    st.markdown("""<div class="settings-card"><p style='font-size:0.88rem;color:#4a5168;line-height:1.7;'>
    Budget is allocated using a multi-factor model:<br><br>
    <b>1. Objective weighting</b> — channels scored by how well they match the campaign objective.<br>
    <b>2. Targetable audience size</b> — captured during the Orchy chat from each platform's estimator. Larger audiences allow more budget before diminishing returns.<br>
    <b>3. Historical cost efficiency</b> — lower historical CPM/CPC = more efficient = higher budget score.<br>
    <b>4. Reach curve</b> — the inflection point controls when diminishing returns kick in. Minimum 5% budget floor per channel.<br>
    <b>5. Data transparency</b> — if no historical data exists for a channel, Sri Lanka industry averages are used and clearly flagged.<br><br>
    <i>In Phase 2, audience sizes and buying rates will be pulled directly from platform APIs.</i>
    </p></div>""", unsafe_allow_html=True)

# Run SQL for new column needed
# ALTER TABLE campaign_plans ADD COLUMN IF NOT EXISTS plan_version text DEFAULT 'V1.0';
# ALTER TABLE campaign_plans ADD COLUMN IF NOT EXISTS budget_split text;
# ALTER TABLE campaign_plans ADD COLUMN IF NOT EXISTS channel_kpi_data text;